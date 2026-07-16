from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.shortcuts import render
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from convocatorias.models import Convocatoria, Postulacion
from exencion.models import Exencion
from formacion.models import ConvocatoriaFormacion, InscripcionFormacion
from registro_audiovisual.models import PersonaHumana, PersonaJuridica


POR_PAGINA = 50


# ============================================================
# HELPERS
# ============================================================

def _admin_url(obj):
    return reverse(
        f"admin:{obj._meta.app_label}_{obj._meta.model_name}_change",
        args=[obj.pk],
    )


def _fmt_fecha(v):
    if v is None:
        return ""
    try:
        return timezone.localtime(v).strftime("%d/%m/%Y %H:%M")
    except Exception:
        pass
    try:
        return v.strftime("%d/%m/%Y")
    except Exception:
        pass
    return str(v)


# B: helper único para filtrar personas (evita duplicación)
def _filtrar_personas(q):
    # A: usa nombre/apellido en lugar de nombre_completo (que es @property)
    humanas_qs = PersonaHumana.objects.select_related("user").all()
    juridicas_qs = PersonaJuridica.objects.select_related("user").all()

    if q:
        humanas_qs = humanas_qs.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(user__email__icontains=q)
        )
        juridicas_qs = juridicas_qs.filter(
            Q(razon_social__icontains=q) | Q(user__email__icontains=q)
        )

    return humanas_qs, juridicas_qs


def _personas_a_rows(humanas_qs, juridicas_qs):
    rows = []
    for ph in humanas_qs:
        rows.append({
            "id":       ph.id,
            "display":  f"{ph.nombre} {ph.apellido}".strip() or "(sin nombre)",
            "tipo":     "Persona humana",
            "email":    getattr(ph.user, "email", "") or "",
            "fecha":    ph.fecha_creacion,
            "url_admin": _admin_url(ph),
        })
    for pj in juridicas_qs:
        rows.append({
            "id":       pj.id,
            "display":  pj.razon_social or "(sin razón social)",
            "tipo":     "Persona jurídica",
            "email":    getattr(pj.user, "email", "") or "",
            "fecha":    pj.fecha_creacion,
            "url_admin": _admin_url(pj),
        })
    rows.sort(key=lambda r: (r["fecha"] is None, r["fecha"]), reverse=True)
    return rows


# ============================================================
# NÓMINA REGISTRO
# ============================================================

@staff_member_required
def nomina_registro(request):
    q = (request.GET.get("q") or "").strip()
    humanas_qs, juridicas_qs = _filtrar_personas(q)
    rows = _personas_a_rows(humanas_qs, juridicas_qs)

    # E: paginación
    paginator = Paginator(rows, POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get("page"))

    return render(request, "backoffice/nomina_registro.html", {
        "q":        q,
        "page_obj": page_obj,
        "total":    paginator.count,
    })


@staff_member_required
def nomina_registro_excel(request):
    q = (request.GET.get("q") or "").strip()
    humanas_qs, juridicas_qs = _filtrar_personas(q)  # B: helper reutilizado

    wb = Workbook()
    wb.remove(wb.active)

    def fmt(v):
        if v is None:
            return ""
        try:
            return timezone.localtime(v).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        try:
            return v.strftime("%d/%m/%Y")
        except Exception:
            pass
        if hasattr(v, "name"):
            return v.name or ""
        return str(v)

    def field_names(model_cls):
        return [getattr(f, "attname", f.name) for f in model_cls._meta.fields]

    def write_sheet(title, qs, model_cls):
        ws = wb.create_sheet(title=title)
        names = field_names(model_cls)
        ws.append(names + ["user_email"])
        for obj in qs:
            row = [fmt(getattr(obj, n, "")) for n in names]
            row.append(getattr(obj.user, "email", "") if obj.user_id else "")
            ws.append(row)
        for i, h in enumerate(names + ["user_email"], 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(h)) + 2, 14), 45)

    write_sheet("Personas Humanas",   humanas_qs,   PersonaHumana)
    write_sheet("Personas Jurídicas", juridicas_qs, PersonaJuridica)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="padron_registro_completo.xlsx"'
    wb.save(response)
    return response


# ============================================================
# CONVOCATORIAS — POSTULACIONES + FORMACIÓN
# ============================================================

@staff_member_required
def backoffice_convocatorias(request):
    q        = (request.GET.get("q") or "").strip()
    conv_id  = request.GET.get("conv_id") or ""
    estado   = request.GET.get("estado") or ""

    # ── POSTULACIONES ────────────────────────────────────────
    post_qs = (
        Postulacion.objects
        .select_related("convocatoria", "user")
        .order_by("-fecha_envio")
    )
    if q:
        post_qs = post_qs.filter(
            Q(nombre_proyecto__icontains=q)
            | Q(user__email__icontains=q)
        )
    if conv_id:
        post_qs = post_qs.filter(convocatoria_id=conv_id)
    if estado:
        post_qs = post_qs.filter(estado=estado)

    # Pre-cargar personas para evitar N+1
    user_ids   = list(post_qs.values_list("user_id", flat=True))
    ph_by_user = {ph.user_id: ph for ph in PersonaHumana.objects.filter(user_id__in=user_ids)}
    pj_by_user = {pj.user_id: pj for pj in PersonaJuridica.objects.filter(user_id__in=user_ids)}

    rows_post = []
    for p in post_qs:
        ph = ph_by_user.get(p.user_id)
        pj = pj_by_user.get(p.user_id)

        if ph:
            presentante = format_html(
                '<a href="{}">{}</a>',
                _admin_url(ph),
                f"{ph.nombre} {ph.apellido}".strip() or str(ph),
            )
        elif pj:
            presentante = format_html(
                '<a href="{}">{}</a>',
                _admin_url(pj),
                pj.razon_social or str(pj),
            )
        else:
            presentante = p.user.email or str(p.user)

        rows_post.append({
            "id":          p.id,
            "presentante": presentante,
            "proyecto":    format_html('<a href="{}">{}</a>', _admin_url(p), p.nombre_proyecto),
            "convocatoria": p.convocatoria.titulo if p.convocatoria_id else "—",
            "linea":       p.convocatoria.get_linea_display() if p.convocatoria_id else "—",
            "fecha_envio": p.fecha_envio,
            "estado":      p.estado,
        })

    # E: paginación postulaciones
    pag_post   = Paginator(rows_post, POR_PAGINA)
    page_post  = pag_post.get_page(request.GET.get("ppost"))

    # ── INSCRIPCIONES FORMACIÓN ──────────────────────────────
    conv_form_id = request.GET.get("conv_form_id") or ""
    estado_form  = request.GET.get("estado_form") or ""

    insc_qs = (
        InscripcionFormacion.objects
        .select_related("convocatoria", "user", "persona_humana", "persona_juridica")
        .order_by("-fecha")
    )
    if q:
        insc_qs = insc_qs.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(user__email__icontains=q)
        )
    if conv_form_id:
        insc_qs = insc_qs.filter(convocatoria_id=conv_form_id)
    if estado_form:
        insc_qs = insc_qs.filter(estado=estado_form)

    rows_form = []
    for i in insc_qs:
        # C: mostrar presentante
        if i.persona_humana:
            ph = i.persona_humana
            presentante = format_html(
                '<a href="{}">{}</a>',
                _admin_url(ph),
                f"{ph.nombre} {ph.apellido}".strip() or str(ph),
            )
        elif i.persona_juridica:
            pj = i.persona_juridica
            presentante = format_html(
                '<a href="{}">{}</a>',
                _admin_url(pj),
                pj.razon_social or str(pj),
            )
        elif i.nombre or i.apellido:
            presentante = f"{i.nombre} {i.apellido}".strip()
        else:
            presentante = i.user.email or str(i.user)

        rows_form.append({
            "id":          format_html('<a href="{}">{}</a>', _admin_url(i), f"#{i.id}"),
            "presentante": presentante,
            "email":       i.user.email or i.email or "—",
            "convocatoria": i.convocatoria.titulo if i.convocatoria_id else "—",
            "fecha":       i.fecha,
            "estado":      i.get_estado_display(),
        })

    pag_form  = Paginator(rows_form, POR_PAGINA)
    page_form = pag_form.get_page(request.GET.get("pform"))

    # ── F: RESUMEN POR ESTADO ────────────────────────────────
    resumen_post = (
        Postulacion.objects
        .values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )
    resumen_form = (
        InscripcionFormacion.objects
        .values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )

    # ── LISTAS PARA FILTROS ──────────────────────────────────
    convocatorias_list      = Convocatoria.objects.order_by("-fecha_inicio").values("id", "titulo")
    conv_formacion_list     = ConvocatoriaFormacion.objects.order_by("-fecha_inicio").values("id", "titulo")
    estados_post_choices    = Postulacion._meta.get_field("estado").choices
    from formacion.models import ESTADOS as ESTADOS_FORM
    estados_form_choices    = ESTADOS_FORM

    return render(request, "backoffice/convocatorias.html", {
        # filtros activos
        "q": q, "conv_id": conv_id, "estado": estado,
        "conv_form_id": conv_form_id, "estado_form": estado_form,
        # datos
        "page_post":  page_post,
        "page_form":  page_form,
        # resúmenes
        "resumen_post": resumen_post,
        "resumen_form": resumen_form,
        # listas para selects
        "convocatorias_list":   convocatorias_list,
        "conv_formacion_list":  conv_formacion_list,
        "estados_post_choices": estados_post_choices,
        "estados_form_choices": estados_form_choices,
    })


# ============================================================
# G: EXENCIONES
# ============================================================

@staff_member_required
def backoffice_exenciones(request):
    q      = (request.GET.get("q") or "").strip()
    estado = (request.GET.get("estado") or "").strip()

    qs = (
        Exencion.objects
        .select_related("user", "persona_humana", "persona_juridica")
        .order_by("-fecha_creacion")
    )
    if q:
        qs = qs.filter(
            Q(nombre_razon_social__icontains=q)
            | Q(cuit__icontains=q)
            | Q(user__email__icontains=q)
        )
    if estado:
        qs = qs.filter(estado=estado)

    rows = []
    for ex in qs:
        rows.append({
            "id":                  ex.id,
            "nombre_razon_social": ex.nombre_razon_social or "—",
            "cuit":                ex.cuit or "—",
            "email":               ex.email or ex.user.email or "—",
            "estado":              ex.get_estado_display(),
            "fecha_creacion":      ex.fecha_creacion,
            "fecha_vencimiento":   ex.fecha_vencimiento,
            "url_admin":           _admin_url(ex),
        })

    paginator = Paginator(rows, POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get("page"))

    resumen = (
        Exencion.objects
        .values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )

    from exencion.models import ESTADOS_EXENCION
    return render(request, "backoffice/exenciones.html", {
        "q":       q,
        "estado":  estado,
        "page_obj": page_obj,
        "total":   paginator.count,
        "resumen": resumen,
        "estados_choices": ESTADOS_EXENCION,
    })
