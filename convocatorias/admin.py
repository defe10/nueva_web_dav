from django.contrib import admin
from django.http import HttpResponse
from django.utils.text import slugify
from django.utils.html import format_html, mark_safe, escape
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone

from . import depuracion

import io
import os
import zipfile

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

from registro_audiovisual.models import PersonaHumana, PersonaJuridica

from .models import (
    Convocatoria,
    MiembroJurado,
    Postulacion,
    DocumentoPostulacion,
    ObservacionAdministrativa,
    AsignacionJuradoConvocatoria,
    Rendicion,
    ConfiguracionPostulacion,
    IntegrantePostulacion,
    DocumentoIntegrante,
    CriterioEvaluacion,
    EvaluacionPostulacion,
    PuntajeCriterio,
)


# ============================================================
#  CONVOCATORIAS
# ============================================================
@admin.register(Convocatoria)
class ConvocatoriaAdmin(admin.ModelAdmin):

    list_display = (
        "titulo",
        "categoria",
        "linea",
        "fecha_inicio",
        "fecha_fin",
        "vigente",
        "orden",
        "url_destino_admin",
    )

    list_filter = (
        "categoria",
        "linea",
        "fecha_inicio",
        "fecha_fin",
    )

    search_fields = (
        "titulo",
        "descripcion_corta",
        "descripcion_larga",
    )

    prepopulated_fields = {"slug": ("titulo",)}
    ordering = ("orden", "-fecha_inicio")

    actions = ["depurar_documentacion_action"]

    def url_destino_admin(self, obj):
        if not obj.url_destino:
            return "—"
        return format_html(
            '<a href="{}" target="_blank" rel="noopener">abrir</a>',
            obj.url_destino
        )
    url_destino_admin.short_description = "URL destino"

    # ==================================================
    # DEPURACIÓN DE DOCUMENTACIÓN (solo superusuarios)
    # ==================================================
    def get_actions(self, request):
        acciones = super().get_actions(request)
        if not request.user.is_superuser:
            acciones.pop("depurar_documentacion_action", None)
        return acciones

    @admin.action(description="Depurar documentación presentada (borrar archivos)")
    def depurar_documentacion_action(self, request, queryset):
        hoy = timezone.localdate()
        cerradas = queryset.filter(fecha_fin__lt=hoy)
        abiertas = queryset.exclude(fecha_fin__lt=hoy)

        # Paso 2: el superusuario ya confirmó en la pantalla intermedia
        if request.POST.get("confirmar_depuracion"):
            incluir = request.POST.get("incluir_ganadores") == "1"
            qs = depuracion.postulaciones_depurables(
                convocatorias=cerradas, incluir_ganadores=incluir,
            )
            resultado = depuracion.ejecutar(qs)
            alcance = "incluyendo ganadoras" if incluir else "ganadoras protegidas"
            messages.success(
                request,
                f"Depuración completada ({alcance}): {resultado['total_docs']} documentos "
                f"borrados, {depuracion.mb(resultado['total_bytes'])} liberados. "
                f"{resultado['marcadas']} postulaciones marcadas como depuradas. "
                "Los datos de las postulaciones se conservan."
            )
            return None

        # Paso 1: pantalla de confirmación con el resumen (nada se borra acá)
        sin_ganadoras = depuracion.postulaciones_depurables(convocatorias=cerradas)
        protegidas = depuracion.ganadoras_protegidas(convocatorias=cerradas)
        resumen = depuracion.resumen(sin_ganadoras)
        resumen_ganadoras = depuracion.resumen(protegidas)

        return render(request, "admin/convocatorias/depurar_confirmacion.html", {
            **self.admin_site.each_context(request),
            "title": "Confirmar depuración de documentación",
            "convocatorias_cerradas": cerradas,
            "convocatorias_abiertas": abiertas,
            "resumen": resumen,
            "resumen_mb": depuracion.mb(resumen["total_bytes"]),
            "resumen_ganadoras": resumen_ganadoras,
            "resumen_ganadoras_mb": depuracion.mb(resumen_ganadoras["total_bytes"]),
            "protegidas": protegidas.count(),
        })


# ============================================================
# INTEGRANTES DEL EQUIPO (inline en Postulación)
# ============================================================
class DocumentoIntegranteInline(admin.TabularInline):
    model = DocumentoIntegrante
    extra = 0
    # Borrado individual de archivos: solo superusuarios (la señal
    # post_delete elimina también el archivo físico).
    can_delete = True
    fields = ("tipo", "estado", "archivo", "fecha_subida")
    readonly_fields = ("archivo", "fecha_subida")

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


class IntegrantePostulacionInline(admin.StackedInline):
    model = IntegrantePostulacion
    extra = 0
    can_delete = False
    fields = ("rol", "persona_humana", "verificado", "documentos_integrante")
    readonly_fields = ("verificado", "documentos_integrante")
    show_change_link = True

    def documentos_integrante(self, obj):
        docs = obj.documentos.all()
        if not docs:
            return "Sin documentación"
        filas = mark_safe("".join(
            "<tr>"
            f"<td>{escape(d.get_tipo_display())}</td>"
            f"<td>{escape(d.get_estado_display())}</td>"
            f"<td><a href='{escape(d.archivo.url)}' target='_blank'>Ver archivo</a></td>"
            "</tr>"
            for d in docs
        ))
        return mark_safe(
            "<table style='font-size:12px'>"
            "<thead><tr><th>Tipo</th><th>Estado</th><th>Archivo</th></tr></thead>"
            f"<tbody>{filas}</tbody>"
            "</table>"
        )
    documentos_integrante.short_description = "Documentación"


# ============================================================
#  DOCUMENTOS DE LA POSTULACIÓN (INLINE)
# ============================================================
class PostulacionDocumentoInline(admin.TabularInline):
    model = DocumentoPostulacion
    extra = 0
    # Borrado individual de archivos: solo superusuarios (la señal
    # post_delete elimina también el archivo físico).
    can_delete = True
    verbose_name = "Documento del proyecto"
    verbose_name_plural = "Documentación del proyecto"
    fields = ("tipo", "estado", "archivo", "fecha_subida")
    readonly_fields = ("archivo", "fecha_subida")
    ordering = ("tipo",)

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser



# ============================================================
#  POSTULACIONES
# ============================================================
@admin.register(Postulacion)
class PostulacionAdmin(admin.ModelAdmin):

    # -------------------------
    # LISTADO
    # -------------------------
    list_display = (
        "id_admin",
        "fecha_envio",
        "presentante",
        "nombre_proyecto",
        "convocatoria",
        "linea_convocatoria",
        "estado",
        
    )

    list_filter = (
        "estado",
        "tipo_proyecto",
        "genero",
        "convocatoria",
        "convocatoria__linea",
        "convocatoria__categoria",
    )

    search_fields = (
        "nombre_proyecto",
        "user__username",
        "user__email",
        "convocatoria__titulo",
    )

    ordering = ("-fecha_envio",)

    actions = [
        "descargar_documentacion_zip",
        "exportar_excel_postulaciones",
        "marcar_admitido",
        "marcar_no_admitido",
        "marcar_evaluacion_jurado",
        "crear_rendicion_para_seleccionados",
        "marcar_seleccionado_y_crear_rendicion",
        "marcar_ganador_y_notificar",
    ]

    # -------------------------
    # DETALLE (base)
    # -------------------------
    readonly_fields = (
        "presentante",
        "convocatoria",
        "fecha_envio",
        "edad",
        "genero_persona",
        "lugar_residencia",
        "documentacion_depurada",
    )

    inlines = [IntegrantePostulacionInline, PostulacionDocumentoInline]

    # ==================================================
    # FIELDSETS DINÁMICOS (línea libre: oculta “Datos del proyecto”)
    # ==================================================
    def get_fieldsets(self, request, obj=None):
        base = (
            ("Datos del presentante", {
                "fields": (
                    "presentante",
                    "fecha_envio",
                    "edad",
                    "genero_persona",
                    "lugar_residencia",
                    "convocatoria",
                    "documentacion_depurada",
                )
            }),
        )

        # Creación (obj=None): mostramos todo
        if obj is None:
            return base + (
                ("Datos del proyecto", {
                    "fields": (
                        "nombre_proyecto",
                        "tipo_proyecto",
                        "genero",
                        "estado",
                        "monto_otorgado",
                    )
                }),
            )

        # Edición: si es línea libre, ocultamos datos del proyecto
        if obj.convocatoria and obj.convocatoria.linea == "libre":
            return base + (
                ("Estado", {
                    "fields": ("estado", "monto_otorgado"),
                    "description": "Línea libre: no requiere datos del proyecto. Se gestiona por documentación."
                }),
            )

        # Normal
        return base + (
            ("Datos del proyecto", {
                "fields": (
                    "nombre_proyecto",
                    "tipo_proyecto",
                    "genero",
                    "estado",
                    "monto_otorgado",
                )
            }),
        )

    # ==================================================
    # READONLY EXTRA (línea libre: bloquea campos del proyecto)
    # ==================================================
    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj and obj.convocatoria and obj.convocatoria.linea == "libre":
            ro.extend(["nombre_proyecto", "tipo_proyecto", "genero"])
        return ro

    # ==================================================
    # QUERYSET CON SELECT_RELATED PARA EVITAR N+1
    # ==================================================
    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("user", "convocatoria", "user__persona_humana", "user__persona_juridica")
        )

    # ==================================================
    # SISTEMA DE EMAILS POR CAMBIO DE ESTADO
    # ==================================================
    _EMAIL_POR_ESTADO = {
        "enviado":           ("convocatorias/email_postulacion_enviada.html", "Tu postulación fue recibida"),
        "admitido":          ("convocatorias/email_admitido.html",            "Tu postulación fue admitida"),
        "no_admitido":       ("convocatorias/email_no_admitido.html",         "Tu postulación no fue admitida"),
        "evaluacion_jurado": ("convocatorias/email_evaluacion_jurado.html",   "Tu postulación está en evaluación"),
        "seleccionado":      ("convocatorias/email_seleccionado.html",        "Tu proyecto fue seleccionado"),
        "no_seleccionado":   ("convocatorias/email_no_seleccionado.html",     "Resultado de tu postulación"),
    }

    def _enviar_email_estado(self, request, postulacion):
        config = self._EMAIL_POR_ESTADO.get(postulacion.estado)
        if not config:
            return
        template, asunto = config
        user = postulacion.user
        if not user or not user.email:
            return
        convocatoria_titulo = postulacion.convocatoria.titulo if postulacion.convocatoria else ""
        panel_url = request.build_absolute_uri(reverse("usuarios:panel_usuario"))
        contexto = {
            "user": user,
            "postulacion": postulacion,
            "convocatoria_titulo": convocatoria_titulo,
            "panel_url": panel_url,
            "anio": timezone.now().year,
        }
        texto = f"{asunto}\n\nConvocatoria: {convocatoria_titulo or '—'}\n\nIngresá al panel: {panel_url}"
        try:
            html = render_to_string(template, contexto)
            email = EmailMultiAlternatives(
                subject=asunto,
                body=texto,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                to=[user.email],
            )
            email.attach_alternative(html, "text/html")
            email.send(fail_silently=False)
            messages.success(request, f"Email enviado a {user.email}: {asunto}.")
        except Exception as e:
            messages.error(request, f"No se pudo enviar el email: {e}")

    def save_model(self, request, obj, form, change):
        estado_anterior = None
        if change and obj.pk:
            estado_anterior = Postulacion.objects.filter(pk=obj.pk).values_list("estado", flat=True).first()
        super().save_model(request, obj, form, change)
        if obj.estado != estado_anterior:
            self._enviar_email_estado(request, obj)

    # ==================================================
    # HELPERS PERSONA (usan el caché del select_related)
    # ==================================================
    def _persona_humana(self, user):
        try:
            return user.persona_humana
        except Exception:
            return None

    def _persona_juridica(self, user):
        try:
            return user.persona_juridica
        except Exception:
            return None

    # ==================================================
    # CAMPOS CALCULADOS
    # ==================================================
    def usuario(self, obj):
        url = reverse("admin:auth_user_change", args=[obj.user_id])
        return format_html('<a href="{}">{}</a>', url, obj.user_id)
    usuario.short_description = "ID"

    def id_admin(self, obj):
        url = reverse("admin:convocatorias_postulacion_change", args=[obj.id])
        return format_html('<a href="{}">{}</a>', url, obj.id)

    id_admin.short_description = "ID"
    id_admin.admin_order_field = "id"



    def linea_convocatoria(self, obj):
        return obj.convocatoria.linea if obj.convocatoria else "—"
    linea_convocatoria.short_description = "Línea"

    def presentante(self, obj):
        ph = self._persona_humana(obj.user)
        pj = self._persona_juridica(obj.user)

        if ph:
            url = reverse("admin:registro_audiovisual_personahumana_change", args=[ph.id])
            return format_html('<a href="{}">{}</a>', url, ph.nombre_completo)

        if pj:
            url = reverse("admin:registro_audiovisual_personajuridica_change", args=[pj.id])
            return format_html('<a href="{}">{}</a>', url, pj.razon_social)

        return "—"
    presentante.short_description = "Presentante"

    def edad(self, obj):
        ph = self._persona_humana(obj.user)
        pj = self._persona_juridica(obj.user)
        if ph and getattr(ph, "edad", None) is not None:
            return ph.edad
        if pj and getattr(pj, "antiguedad", None) is not None:
            return pj.antiguedad
        return "—"
    edad.short_description = "Edad"

    def genero_persona(self, obj):
        ph = self._persona_humana(obj.user)
        if ph and getattr(ph, "genero", None):
            try:
                return ph.get_genero_display()
            except Exception:
                return ph.genero
        return "—"
    genero_persona.short_description = "Género"

    def lugar_residencia(self, obj):
        ph = self._persona_humana(obj.user)
        pj = self._persona_juridica(obj.user)

        if ph:
            try:
                return ph.otro_lugar_residencia if ph.lugar_residencia == "otro" else ph.get_lugar_residencia_display()
            except Exception:
                return "—"

        if pj:
            try:
                return pj.otro_lugar_residencia if pj.lugar_residencia == "otro" else pj.get_lugar_residencia_display()
            except Exception:
                return "—"

        return "—"
    lugar_residencia.short_description = "Lugar de residencia"

    # ==================================================
    # ✅ ACCIONES RENDICIÓN
    # ==================================================
    def crear_rendicion_para_seleccionados(self, request, queryset):
        """
        Crea Rendición SOLO para postulaciones ya 'seleccionado'.
        No modifica el estado de la postulación.
        """
        creadas = 0
        ya_existian = 0
        omitidas = 0

        # optimiza accesos
        queryset = queryset.select_related("user", "convocatoria")

        for p in queryset:
            if p.estado != "seleccionado":
                omitidas += 1
                continue

            rendicion, created = Rendicion.objects.get_or_create(
                postulacion=p,
                defaults={"user": p.user},
            )

            # si existía pero el user no coincide (raro), lo alineamos
            if not created and rendicion.user_id != p.user_id:
                rendicion.user = p.user

            # evento
            try:
                if created:
                    rendicion.add_event("admin", "RENDICION_CREADA", f"Creada desde admin para Postulación {p.id}")
                else:
                    rendicion.add_event("admin", "RENDICION_EXISTENTE", f"Ya existía (admin) para Postulación {p.id}")
                rendicion.save()
            except Exception:
                # si algo raro pasa con JSONField no rompemos la acción
                rendicion.save()

            if created:
                creadas += 1
            else:
                ya_existian += 1

        if creadas:
            messages.success(request, f"✅ Rendiciones creadas: {creadas}.")
        if ya_existian:
            messages.info(request, f"ℹ️ Rendiciones ya existentes: {ya_existian}.")
        if omitidas:
            messages.warning(request, f"⚠️ Omitidas (no estaban en estado 'seleccionado'): {omitidas}.")

    crear_rendicion_para_seleccionados.short_description = "📌 Crear Rendición para SELECCIONADOS"

    def marcar_seleccionado_y_crear_rendicion(self, request, queryset):
        """
        Marca postulaciones como 'seleccionado' y crea Rendición.
        Útil cuando ya está decidida la selección y querés habilitar rendición de una.
        """
        actualizadas = 0
        creadas = 0

        queryset = queryset.select_related("user", "convocatoria")

        for p in queryset:
            # setea seleccionado si no lo estaba
            if p.estado != "seleccionado":
                p.estado = "seleccionado"
                p.save(update_fields=["estado"])
                actualizadas += 1

            rendicion, created = Rendicion.objects.get_or_create(
                postulacion=p,
                defaults={"user": p.user},
            )
            if not created and rendicion.user_id != p.user_id:
                rendicion.user = p.user

            try:
                rendicion.add_event("admin", "SELECCIONADO_Y_RENDICION", f"Postulación {p.id} marcada seleccionado y rendición habilitada")
                rendicion.save()
            except Exception:
                rendicion.save()

            if created:
                creadas += 1

        messages.success(
            request,
            f"✅ Postulaciones marcadas como 'seleccionado': {actualizadas}. Rendiciones creadas: {creadas}."
        )

    marcar_seleccionado_y_crear_rendicion.short_description = "🏆 Marcar como SELECCIONADO + crear Rendición"

    # ==================================================
    # ACCIONES DE ESTADO CON EMAIL
    # ==================================================
    def marcar_ganador_y_notificar(self, request, queryset):
        for p in queryset.select_related("user", "convocatoria"):
            if p.estado != "seleccionado":
                p.estado = "seleccionado"
                p.save(update_fields=["estado"])
            self._enviar_email_estado(request, p)
        self.message_user(request, f"{queryset.count()} proyecto(s) marcado(s) como seleccionado/s y notificado/s.")

    marcar_ganador_y_notificar.short_description = "✉️ Marcar como SELECCIONADO y enviar email"

    def marcar_admitido(self, request, queryset):
        for p in queryset.select_related("user", "convocatoria"):
            if p.estado != "admitido":
                p.estado = "admitido"
                p.save(update_fields=["estado"])
                self._enviar_email_estado(request, p)
        self.message_user(request, f"{queryset.count()} postulación/es marcada/s como admitida/s.")

    marcar_admitido.short_description = "✅ Marcar como ADMITIDO y notificar"

    def marcar_no_admitido(self, request, queryset):
        for p in queryset.select_related("user", "convocatoria"):
            if p.estado != "no_admitido":
                p.estado = "no_admitido"
                p.save(update_fields=["estado"])
                self._enviar_email_estado(request, p)
        self.message_user(request, f"{queryset.count()} postulación/es marcada/s como no admitida/s.")

    marcar_no_admitido.short_description = "❌ Marcar como NO ADMITIDO y notificar"

    def marcar_evaluacion_jurado(self, request, queryset):
        for p in queryset.select_related("user", "convocatoria"):
            if p.estado != "evaluacion_jurado":
                p.estado = "evaluacion_jurado"
                p.save(update_fields=["estado"])
                self._enviar_email_estado(request, p)
        self.message_user(request, f"{queryset.count()} postulación/es enviada/s a evaluación por jurado.")

    marcar_evaluacion_jurado.short_description = "⚖️ Enviar a EVALUACIÓN POR JURADO y notificar"

    # ==================================================
    # ACCIÓN ZIP
    # ==================================================
    def descargar_documentacion_zip(self, request, queryset):
        if queryset.count() == 1:
            return self._zip_para_una_postulacion(queryset.first())

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in queryset:
                zip_bytes, zip_name = self._bytes_zip_para_una_postulacion(p)
                zf.writestr(zip_name, zip_bytes.getvalue())

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="documentacion_postulaciones.zip"'
        return response

    descargar_documentacion_zip.short_description = "📦 Descargar documentación (ZIP)"

    def _agregar_archivo_zip(self, zf, file_field, carpeta, nombre):
        """Agrega el archivo al ZIP. Devuelve True si se pudo agregar."""
        if not file_field:
            return False
        try:
            path = getattr(file_field, "path", None)
            if path and os.path.exists(path):
                zf.write(path, f"{carpeta}/{nombre}{os.path.splitext(path)[1]}")
            else:
                with file_field.open("rb") as f:
                    zf.writestr(
                        f"{carpeta}/{nombre}{os.path.splitext(file_field.name)[1]}",
                        f.read()
                    )
            return True
        except Exception:
            return False

    def _safe_slug(self, p):
        base = p.nombre_proyecto or f"postulacion_{p.id}"
        return slugify(base)[:40] or f"postulacion_{p.id}"

    def _bytes_zip_para_una_postulacion(self, p):
        buffer = io.BytesIO()
        name = f"postulacion_{p.id}_{self._safe_slug(p)}.zip"
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            self._armar_zip(zf, p)
        buffer.seek(0)
        return buffer, name

    def _zip_para_una_postulacion(self, p):
        buffer = io.BytesIO()
        name = f"postulacion_{p.id}_{self._safe_slug(p)}.zip"
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            self._armar_zip(zf, p)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{name}"'
        return response

    def _armar_zip(self, zf, p):
        zf.writestr(
            "README.txt",
            f"Postulación ID: {p.id}\n"
            f"Proyecto: {p.nombre_proyecto or '(Sin título)'}\n"
            f"Usuario: {p.user.username}\n"
            f"Convocatoria: {p.convocatoria.titulo if p.convocatoria else ''}\n"
        )
        faltantes = []
        for doc in p.documentos.all():
            # id incluido para que documentos del mismo tipo no se pisen en el ZIP
            agregado = self._agregar_archivo_zip(
                zf,
                doc.archivo,
                "documentacion",
                slugify(f"{doc.get_tipo_display()}-{doc.id}")
            )
            if not agregado:
                faltantes.append(
                    f"- {doc.get_tipo_display()} (id {doc.id}): "
                    f"{doc.archivo.name if doc.archivo else 'sin archivo'}"
                )
        for integrante in p.integrantes.all():
            carpeta = f"equipo/{integrante.rol.lower()}_{slugify(integrante.nombre_busqueda or str(integrante.id))}"
            for doc in integrante.documentos.all():
                agregado = self._agregar_archivo_zip(
                    zf,
                    doc.archivo,
                    carpeta,
                    slugify(f"{doc.get_tipo_display()}-{doc.id}"),
                )
                if not agregado:
                    faltantes.append(
                        f"- {carpeta} · {doc.get_tipo_display()} (id {doc.id}): "
                        f"{doc.archivo.name if doc.archivo else 'sin archivo'}"
                    )
        if faltantes:
            zf.writestr(
                "FALTANTES.txt",
                "Documentos que NO se pudieron incluir en este ZIP:\n"
                + "\n".join(faltantes)
                + "\n"
            )

    # ==================================================
    # EXPORTAR EXCEL
    # ==================================================
    def exportar_excel_postulaciones(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Postulaciones"

        headers = [
            "Fecha postulación",
            "Usuario",
            "Presentante",
            "Edad",
            "Género (persona)",
            "Lugar de residencia",
            "Convocatoria",
            "Línea",
            "Nombre del proyecto",
            "Tipo de proyecto",
            "Género (proyecto)",
            "Estado",
        ]
        ws.append(headers)

        queryset = queryset.select_related("user", "convocatoria", "user__persona_humana", "user__persona_juridica")

        for p in queryset:
            ws.append([
                p.fecha_envio.strftime("%d/%m/%Y %H:%M") if p.fecha_envio else "",
                p.user.username,
                self._presentante_texto(p.user),
                self._edad_texto(p.user),
                self._genero_persona_texto(p.user),
                self._lugar_residencia_texto(p.user),
                p.convocatoria.titulo if p.convocatoria else "",
                p.convocatoria.linea if p.convocatoria else "",
                p.nombre_proyecto or "",
                p.get_tipo_proyecto_display() if p.tipo_proyecto else "",
                p.get_genero_display() if p.genero else "",
                p.get_estado_display() if p.estado else "",
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="postulaciones.xlsx"'
        wb.save(response)
        return response

    exportar_excel_postulaciones.short_description = "📤 Exportar seleccionadas a Excel (.xlsx)"

    def _presentante_texto(self, user):
        ph = self._persona_humana(user)
        pj = self._persona_juridica(user)
        if ph:
            return ph.nombre_completo
        if pj:
            return pj.razon_social
        return ""

    def _edad_texto(self, user):
        ph = self._persona_humana(user)
        pj = self._persona_juridica(user)
        if ph and getattr(ph, "edad", None) is not None:
            return ph.edad
        if pj and getattr(pj, "antiguedad", None) is not None:
            return pj.antiguedad
        return ""

    def _genero_persona_texto(self, user):
        ph = self._persona_humana(user)
        if ph and getattr(ph, "genero", None):
            try:
                return ph.get_genero_display()
            except Exception:
                return ph.genero
        return ""

    def _lugar_residencia_texto(self, user):
        ph = self._persona_humana(user)
        pj = self._persona_juridica(user)
        if ph:
            try:
                return ph.otro_lugar_residencia if ph.lugar_residencia == "otro" else ph.get_lugar_residencia_display()
            except Exception:
                return ""
        if pj:
            try:
                return pj.otro_lugar_residencia if pj.lugar_residencia == "otro" else pj.get_lugar_residencia_display()
            except Exception:
                return ""
        return ""


# ============================================================
# OBSERVACIONES ADMINISTRATIVAS
# ============================================================
@admin.register(ObservacionAdministrativa)
class ObservacionAdministrativaAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "proyecto_link",
        "presentante_link",
        "tipo_documento",
        "descripcion",
        "subsanada",
        "fecha_creacion",
    )
    list_filter = ("subsanada", "tipo_documento")
    search_fields = ("descripcion", "postulacion__nombre_proyecto", "postulacion__id")

    # -------------------------
    # LINKS EN LISTADO
    # -------------------------
    def proyecto_link(self, obj):
        """
        Proyecto (link a Postulación).
        Si no hay nombre_proyecto, muestra (Sin título).
        """
        p = getattr(obj, "postulacion", None)
        if not p:
            return "—"

        url = reverse("admin:convocatorias_postulacion_change", args=[p.id])
        nombre = getattr(p, "nombre_proyecto", None) or "(Sin título)"
        return format_html('<a href="{}">{}</a>', url, nombre)

    proyecto_link.short_description = "Proyecto"
    proyecto_link.admin_order_field = "postulacion__nombre_proyecto"

    def presentante_link(self, obj):
        """
        Presentante (link al Registro: PersonaHumana o PersonaJuridica).
        """
        p = getattr(obj, "postulacion", None)
        if not p or not getattr(p, "user", None):
            return "—"

        user = p.user
        ph = PersonaHumana.objects.filter(user=user).first()
        pj = PersonaJuridica.objects.filter(user=user).first()

        if ph:
            url = reverse("admin:registro_audiovisual_personahumana_change", args=[ph.id])
            return format_html('<a href="{}">{}</a>', url, ph.nombre_completo)

        if pj:
            url = reverse("admin:registro_audiovisual_personajuridica_change", args=[pj.id])
            return format_html('<a href="{}">{}</a>', url, pj.razon_social)

        # fallback: si no tiene registro, mostramos username sin link
        return user.username or "—"

    presentante_link.short_description = "Presentante"

    def save_model(self, request, obj, form, change):
        """
        Envía email cuando:
        - Se crea una observación pendiente (subsanada=False), o
        - Se edita y sigue pendiente, y cambió algo relevante (tipo/descripcion/subsanada)
        """
        es_nueva = obj.pk is None

        anterior = None
        if not es_nueva:
            anterior = ObservacionAdministrativa.objects.filter(pk=obj.pk).first()

        super().save_model(request, obj, form, change)

        # Solo avisamos si está pendiente
        if obj.subsanada:
            return

        # Si es edición: solo mandar si cambió algo relevante
        if anterior and (
            anterior.subsanada == obj.subsanada
            and anterior.tipo_documento == obj.tipo_documento
            and (anterior.descripcion or "").strip() == (obj.descripcion or "").strip()
        ):
            return

        postulacion = obj.postulacion
        user = getattr(postulacion, "user", None)
        if not user:
            return

        destinatario = user.email
        if not destinatario:
            messages.warning(request, "No se envió email: el usuario no tiene email cargado.")
            return

        convocatoria_titulo = ""
        if getattr(postulacion, "convocatoria", None):
            convocatoria_titulo = postulacion.convocatoria.titulo or ""

        panel_url = None
        try:
            panel_url = request.build_absolute_uri(reverse("usuarios:panel_usuario"))
        except Exception:
            panel_url = None

        if not panel_url:
            try:
                panel_url = request.build_absolute_uri(reverse("convocatorias:convocatorias_home"))
            except Exception:
                panel_url = None

        if not panel_url:
            panel_url = request.build_absolute_uri("/usuarios/panel/")

        asunto = "Subsanación de documentación requerida"

        contexto = {
            "user": user,
            "postulacion": postulacion,
            "convocatoria_titulo": convocatoria_titulo,
            "observacion": obj,
            "panel_url": panel_url,
            "anio": timezone.now().year,
        }

        texto = (
            "Tenés una subsanación de documentación pendiente.\n\n"
            f"Convocatoria: {convocatoria_titulo or '—'}\n"
            f"Documento: {getattr(obj, 'get_tipo_documento_display', lambda: obj.tipo_documento)()}\n"
            f"Detalle: {obj.descripcion}\n\n"
            + (f"Ingresá al panel para subsanar: {panel_url}\n" if panel_url else "")
        )

        try:
            try:
                html = render_to_string(
                    "convocatorias/subsanacion_documentacion_email.html",
                    contexto
                )
            except Exception:
                html = None

            email = EmailMultiAlternatives(
                subject=asunto,
                body=texto,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                to=[destinatario],
            )
            if html:
                email.attach_alternative(html, "text/html")
            email.send(fail_silently=False)

            messages.success(request, f"Email de subsanación enviado a {destinatario}.")
        except Exception as e:
            messages.error(request, f"No se pudo enviar el email de subsanación: {e}")



# ============================================================
# ASIGNACIÓN JURADO ⇄ CONVOCATORIA
# ============================================================
@admin.register(AsignacionJuradoConvocatoria)
class AsignacionJuradoConvocatoriaAdmin(admin.ModelAdmin):
    list_display  = ("jurado", "convocatoria", "doble_ciego", "fecha_asignacion")
    list_filter   = ("convocatoria", "doble_ciego")
    list_editable = ("doble_ciego",)
    search_fields = ("jurado__username", "convocatoria__titulo")


# ============================================================
# RENDICIÓN
# ============================================================
@admin.register(Rendicion)
class RendicionAdmin(admin.ModelAdmin):

    list_display = (
        "id",
        "presentante",
        "convocatoria_link",
        "nombre_proyecto",
        "estado_postulacion",
        "fecha_creacion",
    )

    ordering = ("-fecha_creacion",)

    # =========================
    # LINKS Y CAMPOS CALCULADOS
    # =========================

    def nombre_proyecto(self, obj):
        postulacion = obj.postulacion
        if not postulacion:
            return "—"

        url = reverse("admin:convocatorias_postulacion_change", args=[postulacion.id])
        nombre = postulacion.nombre_proyecto or "(Sin título)"
        return format_html('<a href="{}">{}</a>', url, nombre)

    nombre_proyecto.short_description = "Proyecto"

    def convocatoria_link(self, obj):
        postulacion = obj.postulacion
        if not postulacion or not postulacion.convocatoria:
            return "—"

        conv = postulacion.convocatoria
        url = reverse("admin:convocatorias_convocatoria_change", args=[conv.id])
        return format_html('<a href="{}">{}</a>', url, conv.titulo)

    convocatoria_link.short_description = "Convocatoria"

    def presentante(self, obj):
        postulacion = obj.postulacion
        if not postulacion:
            return "—"

        user = postulacion.user
        ph = PersonaHumana.objects.filter(user=user).first()
        pj = PersonaJuridica.objects.filter(user=user).first()

        if ph:
            url = reverse("admin:registro_audiovisual_personahumana_change", args=[ph.id])
            return format_html('<a href="{}">{}</a>', url, ph.nombre_completo)

        if pj:
            url = reverse("admin:registro_audiovisual_personajuridica_change", args=[pj.id])
            return format_html('<a href="{}">{}</a>', url, pj.razon_social)

        return user.username

    presentante.short_description = "Presentante"

    def estado_postulacion(self, obj):
        if not obj.postulacion:
            return "—"
        return obj.postulacion.get_estado_display()

    estado_postulacion.short_description = "Estado"

    fieldsets = (
        ("Postulación", {
            "fields": ("postulacion", "user", "estado", "fecha_envio"),
        }),
        ("Documentación digital", {
            "fields": ("planilla_xlsx", "link_documentacion", "observaciones_usuario"),
        }),
        ("Impacto económico", {
            "description": "Montos totales y cantidades por categoría extraídos de la planilla de rendición.",
            "fields": (
                ("honorarios_tecnicos",     "honorarios_tecnicos_cantidad"),
                ("honorarios_elenco",       "honorarios_elenco_cantidad"),
                ("otros_honorarios",        "otros_honorarios_cantidad"),
                ("insumos",                 "insumos_cantidad"),
                ("servicios_audiovisuales", "servicios_audiovisuales_cantidad"),
                ("servicios_logistica",     "servicios_logistica_cantidad"),
            ),
        }),
        ("Revisión administrativa", {
            "fields": ("observaciones_admin", "fecha_ultima_revision", "fecha_aprobacion"),
        }),
        ("Estado físico", {
            "classes": ("collapse",),
            "fields": ("fisico_estado", "fisico_fecha_recepcion", "fisico_observaciones"),
        }),
    )

    def save_model(self, request, obj, form, change):
        if obj.estado == "APROBADO" and not obj.fecha_aprobacion:
            obj.fecha_aprobacion = timezone.now().date()
            obj.add_event("admin", "APROBADO", f"Aprobada por {request.user.username}")
        super().save_model(request, obj, form, change)




# ============================================================
# CONFIGURACIÓN DE POSTULACIÓN (inline en Convocatoria)
# ============================================================
class ConfiguracionPostulacionInline(admin.StackedInline):
    model = ConfiguracionPostulacion
    can_delete = False
    verbose_name = "Configuración de postulación"
    verbose_name_plural = "Configuración de postulación"
    fieldsets = (
        ("Equipo", {
            "fields": (
                "tipo_postulante",
                "requiere_productor_responsable",
                "requiere_director",
                "director_puede_coincidir",
                "requiere_guionista",
                "requiere_realizador",
                "requiere_cbu",
            )
        }),
        ("Proyecto — texto", {
            "fields": (
                "requiere_sinopsis",
                "requiere_link_pitch",
            )
        }),
        ("Proyecto — documentos", {
            "description": "Activá los documentos que se mostrarán en el formulario. Todos son opcionales para el postulante.",
            "fields": (
                "mostrar_guion",
                "mostrar_dossier",
                "mostrar_material_adicional",
                "mostrar_planilla_oficial",
                "mostrar_dnda",
                "mostrar_autorizacion_derechos",
                "mostrar_nota_intencion",
                "mostrar_carta_intencion",
                "mostrar_constancia_invitacion",
                "mostrar_documentacion",
            )
        }),
        ("Planilla oficial", {
            "description": "Archivo xlsx que el presentante descarga, completa offline y sube al postularse.",
            "fields": ("planilla_archivo",),
        }),
    )


# ============================================================
# REGISTRO ADMIN NUEVOS MODELOS
# ============================================================
@admin.register(ConfiguracionPostulacion)
class ConfiguracionPostulacionAdmin(admin.ModelAdmin):
    list_display  = ("convocatoria", "tipo_postulante", "requiere_productor_responsable", "requiere_director", "requiere_guionista", "requiere_realizador")
    list_filter   = ("tipo_postulante", "requiere_productor_responsable", "requiere_director", "requiere_guionista")
    search_fields = ("convocatoria__titulo",)


@admin.register(IntegrantePostulacion)
class IntegrantePostulacionAdmin(admin.ModelAdmin):
    list_display  = ("postulacion", "rol", "nombre_busqueda", "persona_humana", "verificado")
    list_filter   = ("rol", "verificado")
    search_fields = ("nombre_busqueda", "persona_humana__nombre_completo", "postulacion__nombre_proyecto")
    inlines       = [DocumentoIntegranteInline]


# ============================================================
# MIEMBROS DEL JURADO (inline en Convocatoria)
# ============================================================
class MiembroJuradoInline(admin.TabularInline):
    model = MiembroJurado
    extra = 1
    fields = ("orden", "nombre", "foto", "bio")


class CriterioEvaluacionInline(admin.TabularInline):
    model = CriterioEvaluacion
    extra = 1
    fields = ("orden", "nombre", "puntaje_maximo")
    ordering = ("orden",)
    verbose_name = "Criterio de evaluación"
    verbose_name_plural = "Criterios de evaluación"


# ============================================================
# CONVOCATORIA — agregar inlines
# ============================================================
ConvocatoriaAdmin.inlines = [ConfiguracionPostulacionInline, MiembroJuradoInline, CriterioEvaluacionInline]


# @admin.register(DocumentoPostulacion)
# class DocumentoPostulacionAdmin(admin.ModelAdmin):
#     list_display = (
#         "postulacion",
#         "tipo",
#         "subtipo_subsanado",
#         "estado",
#         "fecha_subida",
#     )
#     list_filter = ("tipo", "subtipo_subsanado", "estado")
#     search_fields = (
#         "postulacion__id",
#         "postulacion__nombre_proyecto",
#         "postulacion__user__username",
#         "postulacion__user__email",
#     )
#     ordering = ("-fecha_subida",)

#     def save_model(self, request, obj, form, change):
#         # Si no es SUBSANADO, el subtipo no debe quedar seteado
#         if obj.tipo != "SUBSANADO":
#             obj.subtipo_subsanado = None

#         # Si es SUBSANADO, forzamos que tenga subtipo (para que jurado vea lo correcto)
#         if obj.tipo == "SUBSANADO" and not obj.subtipo_subsanado:
#             messages.error(
#                 request,
#                 "⚠️ Si el tipo es SUBSANADO, debés seleccionar Subtipo (PROYECTO o ADMIN)."
#             )
#             return  # corta el guardado

#         super().save_model(request, obj, form, change)


# ============================================================
#  EVALUACIÓN DEL COMITÉ
# ============================================================

class PuntajeCriterioInline(admin.TabularInline):
    model = PuntajeCriterio
    extra = 0
    fields = ("criterio", "puntaje")
    readonly_fields = ()

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "criterio" and hasattr(request, "_evaluacion_obj"):
            kwargs["queryset"] = CriterioEvaluacion.objects.filter(
                convocatoria=request._evaluacion_obj.postulacion.convocatoria
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class CriterioEvaluacionAdmin(admin.ModelAdmin):
    list_display  = ("convocatoria", "orden", "nombre", "puntaje_maximo")
    list_filter   = ("convocatoria",)
    ordering      = ("convocatoria", "orden")


@admin.register(EvaluacionPostulacion)
class EvaluacionPostulacionAdmin(admin.ModelAdmin):
    list_display  = ("postulacion", "puntaje_total", "no_puntuar", "fecha_modificacion", "ultima_edicion_por")
    list_filter   = ("postulacion__convocatoria", "no_puntuar")
    search_fields = ("postulacion__nombre_proyecto",)
    readonly_fields = ("ultima_edicion_por", "fecha_modificacion", "puntaje_total")
    inlines       = [PuntajeCriterioInline]
    actions       = ["marcar_ganador_y_notificar", "descargar_acta_jurado"]

    def get_form(self, request, obj=None, **kwargs):
        if obj:
            request._evaluacion_obj = obj
        return super().get_form(request, obj, **kwargs)

    def save_model(self, request, obj, form, change):
        obj.ultima_edicion_por = request.user
        super().save_model(request, obj, form, change)

    def puntaje_total(self, obj):
        return obj.puntaje_total
    puntaje_total.short_description = "Puntaje total"

    def marcar_ganador_y_notificar(self, request, queryset):
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        from django.urls import reverse

        EMAIL_TEMPLATE = "convocatorias/email_seleccionado.html"
        ASUNTO = "Tu proyecto fue seleccionado"

        for ev in queryset.select_related("postulacion__user", "postulacion__convocatoria"):
            p = ev.postulacion
            if p.estado != "seleccionado":
                p.estado = "seleccionado"
                p.save(update_fields=["estado"])

            user = p.user
            if not user or not user.email:
                continue

            convocatoria_titulo = p.convocatoria.titulo if p.convocatoria else ""
            panel_url = request.build_absolute_uri(reverse("usuarios:panel_usuario"))
            contexto = {
                "user": user,
                "postulacion": p,
                "convocatoria_titulo": convocatoria_titulo,
                "panel_url": panel_url,
                "anio": timezone.now().year,
            }
            texto = f"{ASUNTO}\n\nConvocatoria: {convocatoria_titulo}\n\nIngresá al panel: {panel_url}"
            try:
                html = render_to_string(EMAIL_TEMPLATE, contexto)
                email = EmailMultiAlternatives(
                    subject=ASUNTO,
                    body=texto,
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                    to=[user.email],
                )
                email.attach_alternative(html, "text/html")
                email.send(fail_silently=False)
                messages.success(request, f"Email enviado a {user.email}.")
            except Exception as e:
                messages.error(request, f"Error enviando email a {user.email}: {e}")

    marcar_ganador_y_notificar.short_description = "Marcar como ganador y enviar email"

    def descargar_acta_jurado(self, request, queryset):
        convocatorias = queryset.values_list(
            "postulacion__convocatoria", flat=True
        ).distinct()
        if convocatorias.count() > 1:
            self.message_user(
                request,
                "Seleccioná evaluaciones de una sola convocatoria para generar el acta.",
                level=messages.ERROR,
            )
            return

        from .models import MiembroJurado
        convocatoria = queryset.first().postulacion.convocatoria
        jurados = MiembroJurado.objects.filter(convocatoria=convocatoria).order_by("orden")

        evaluaciones = list(
            queryset
            .select_related("postulacion")
            .prefetch_related("puntajes")
        )
        evaluaciones.sort(key=lambda e: (e.no_puntuar, -(e.puntaje_total or 0)))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=3*cm, rightMargin=3*cm,
            topMargin=2.5*cm, bottomMargin=2.5*cm,
        )

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
            fontSize=14, spaceAfter=6, alignment=TA_CENTER)
        subtitulo_style = ParagraphStyle("subtitulo", parent=styles["Normal"],
            fontSize=11, spaceAfter=4, alignment=TA_CENTER)
        cuerpo_style = ParagraphStyle("cuerpo", parent=styles["Normal"],
            fontSize=10, spaceAfter=6, leading=14, alignment=TA_JUSTIFY)
        proyecto_style = ParagraphStyle("proyecto", parent=styles["Normal"],
            fontSize=11, spaceBefore=10, spaceAfter=2,
            textColor=colors.HexColor("#1a1a2e"), fontName="Helvetica-Bold")
        fundamentacion_style = ParagraphStyle("fundamentacion", parent=styles["Normal"],
            fontSize=9, leading=13, leftIndent=10, alignment=TA_JUSTIFY,
            textColor=colors.HexColor("#444444"))
        firma_style = ParagraphStyle("firma", parent=styles["Normal"],
            fontSize=9, alignment=TA_CENTER)

        from django.utils.formats import date_format
        fecha_hoy = timezone.localtime(timezone.now()).strftime("%-d de %B de %Y")
        nombres_jurados = ", ".join(j.nombre for j in jurados) if jurados else "el comité evaluador"

        LOGO_PATH = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "static", "exencion", "img", "sec_cultura_color.png"
        )

        story = []

        # Logo
        if os.path.exists(LOGO_PATH):
            logo = Image(LOGO_PATH, hAlign="LEFT")
            logo.drawWidth = 7*cm
            logo.drawHeight = logo.drawWidth * (150 / 1040)
            story.append(logo)
            story.append(Spacer(1, 0.8*cm))

        encabezado_style = ParagraphStyle(
            "encabezado", parent=styles["Normal"],
            fontSize=13, fontName="Helvetica-Bold",
            alignment=TA_CENTER, leading=16, spaceAfter=2,
        )
        subencabezado_style = ParagraphStyle(
            "subencabezado", parent=styles["Normal"],
            fontSize=10, alignment=TA_CENTER, leading=13, spaceAfter=0,
        )
        story.append(Paragraph("Dirección de Audiovisuales", encabezado_style))
        story.append(Paragraph("Secretaría de Cultura de la Provincia de Salta", subencabezado_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.black))
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("ACTA DE EVALUACIÓN — COMITÉ DE JURADO", titulo_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"Convocatoria: <b>{convocatoria.titulo}</b>", cuerpo_style))
        story.append(Paragraph(f"Fecha: {fecha_hoy}", cuerpo_style))
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(
            f"Los integrantes del Comité Evaluador, <b>{nombres_jurados}</b>, reunidos en el marco "
            f"de la convocatoria <b>{convocatoria.titulo}</b>, habiendo analizado los proyectos postulados "
            f"conforme a los criterios de evaluación establecidos en las Bases y Condiciones, proceden "
            f"a establecer el siguiente orden de mérito:",
            cuerpo_style,
        ))
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        story.append(Spacer(1, 0.4*cm))

        label_style = ParagraphStyle(
            "label", parent=styles["Normal"],
            fontSize=9, leading=13, leftIndent=10,
            textColor=colors.HexColor("#444444"),
        )

        puesto = 1
        for ev in evaluaciones:
            p = ev.postulacion
            user = p.user
            ph = PersonaHumana.objects.filter(user=user).first()
            pj = PersonaJuridica.objects.filter(user=user).first()
            nombre_presentante = (
                ph.nombre_completo if ph else
                pj.razon_social if pj else
                user.get_full_name() or user.username
            )

            if ev.no_puntuar:
                titulo_linea = f"— {p.nombre_proyecto}  [No puntuado]"
            else:
                puntaje = ev.puntaje_total if ev.puntaje_total is not None else "—"
                titulo_linea = f"{puesto}. {p.nombre_proyecto}"
                puesto += 1

            story.append(Paragraph(titulo_linea, proyecto_style))
            story.append(Paragraph(f"<b>Presentante:</b> {nombre_presentante}", label_style))
            if not ev.no_puntuar:
                story.append(Paragraph(f"<b>Puntaje obtenido:</b> {puntaje} pts", label_style))
            if ev.fundamentacion:
                story.append(Spacer(1, 0.15*cm))
                story.append(Paragraph(f"<b>Fundamentación:</b> {ev.fundamentacion}", fundamentacion_style))
            story.append(Spacer(1, 0.4*cm))

        story.append(Spacer(1, 1.5*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        story.append(Spacer(1, 1*cm))

        if jurados:
            ancho_col = (A4[0] - 6*cm) / max(len(jurados), 1)
            firma_data = [[Paragraph(j.nombre, firma_style) for j in jurados]]
            tabla_firmas = Table(firma_data, colWidths=[ancho_col] * len(jurados))
            tabla_firmas.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.black),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(tabla_firmas)

        doc.build(story)
        buffer.seek(0)

        nombre_archivo = f"acta_jurado_{slugify(convocatoria.titulo)}.pdf"
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
        return response

    descargar_acta_jurado.short_description = "Descargar acta de jurado (PDF)"

