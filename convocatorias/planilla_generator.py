"""
Genera planillas oficiales IDEA pre-completadas a partir de las plantillas xlsx.
"""
import io
import os

from django.conf import settings
from openpyxl import load_workbook

PLANILLAS_DIR = getattr(
    settings,
    "PLANILLAS_OFICIALES_DIR",
    "/Users/hasguell/Library/CloudStorage/GoogleDrive-federicocasoni@gmail.com"
    "/Mi unidad/DAV/PROYECTOS 2026/FOMENTO/PLAN IDEA 2026/Planillas oficiales",
)

PLANILLA_FILES = {
    "01": "01_ Planilla Oficial_Cortometraje ficcion.xlsx",
    "02": "02_ Planilla Oficial_Cortometraje coproducción.xlsx",
    "03": "03_ Planilla Oficial_Desarrollo.xlsx",
    "04": "04_ Planilla Oficial_Lab corto.xlsx",
    "05": "05_ Planilla Oficial_Cine en comunidad.xlsx",
    "06": "06_ Planilla Oficial_Apoyo rodaje.xlsx",
    "07": "07_ Planilla Oficial_Apoyo finalización.xlsx",
    "08": "08_ Planilla Oficial_Cash rebate.xlsx",
}


def generar_planilla_postulacion(postulacion) -> tuple[bytes, str]:
    """
    Devuelve (bytes_del_xlsx, nombre_de_archivo).
    Toma la plantilla correspondiente a la convocatoria, pre-completa los datos
    del presentante y del proyecto, y la retorna como bytes lista para descargar.
    """
    config = getattr(postulacion.convocatoria, "configuracion", None)
    template_key = config.planilla_template if config else ""

    if not template_key or template_key not in PLANILLA_FILES:
        raise ValueError(
            f"La convocatoria '{postulacion.convocatoria.titulo}' "
            "no tiene una planilla oficial configurada."
        )

    template_path = os.path.join(PLANILLAS_DIR, PLANILLA_FILES[template_key])
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró el archivo de plantilla: {template_path}"
        )

    wb = load_workbook(template_path)
    ws = wb["01_Inicio"]

    # Nombre del presentante
    persona = _get_persona(postulacion)
    ws["B4"] = persona or postulacion.user.get_full_name() or postulacion.user.username

    # Nombre del proyecto
    ws["B5"] = postulacion.nombre_proyecto or ""

    # Parámetros financieros (desde la config de la convocatoria)
    if config:
        ws["B8"] = config.beneficio_monto or 0
        ws["B9"] = config.beneficio_adicional or 0
        ws["B10"] = float(config.aporte_propio_pct or 0.10)

    nombre_archivo = (
        f"planilla_{template_key}_"
        f"{(postulacion.nombre_proyecto or 'proyecto').replace(' ', '_')}.xlsx"
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), nombre_archivo


def _get_persona(postulacion):
    """Retorna el nombre completo del presentante desde el registro audiovisual."""
    user = postulacion.user
    try:
        return user.persona_humana.nombre_completo
    except Exception:
        pass
    try:
        return user.persona_juridica.razon_social
    except Exception:
        pass
    return None
