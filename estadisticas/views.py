from datetime import date
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from convocatorias.models import Postulacion, Convocatoria, Rendicion
from registro_audiovisual.models import PersonaHumana, GENERO_CHOICES


ESTADOS_GANADOR = {"seleccionado", "finalizado"}

RANGOS_ETARIOS = [
    ("18-25", 18, 25),
    ("26-35", 26, 35),
    ("36-45", 36, 45),
    ("46-55", 46, 55),
    ("56+",   56, 999),
]


def _edad(persona):
    if not persona or not persona.fecha_nacimiento:
        return None
    hoy = date.today()
    fn = persona.fecha_nacimiento
    return hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))


def _rango(edad):
    if edad is None:
        return "Sin dato"
    for label, lo, hi in RANGOS_ETARIOS:
        if lo <= edad <= hi:
            return label
    return "Sin dato"


def _postulaciones_qs(filtros):
    qs = Postulacion.objects.exclude(estado="borrador").select_related(
        "convocatoria", "user"
    )
    if filtros.get("conv"):
        qs = qs.filter(convocatoria_id=filtros["conv"])
    if filtros.get("anio"):
        qs = qs.filter(fecha_envio__year=filtros["anio"])
    if filtros.get("tipo"):
        qs = qs.filter(tipo_proyecto=filtros["tipo"])
    if filtros.get("solo_ganadores"):
        qs = qs.filter(estado__in=ESTADOS_GANADOR)
    return qs


TIPO_MAP   = {v: l for v, l in Postulacion.TIPO_PROYECTO if v}
GENERO_MAP = {v: l for v, l in Postulacion.GENERO if v}
ESTADO_MAP = {v: l for v, l in Postulacion.ESTADOS if v}
GENERO_PH_MAP = dict(GENERO_CHOICES)


def _agrupar(qs):
    por_tipo = [
        {"label": TIPO_MAP.get(r["tipo_proyecto"], r["tipo_proyecto"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("tipo_proyecto").annotate(total=Count("id")).order_by("-total")
    ]
    por_genero_proy = [
        {"label": GENERO_MAP.get(r["genero"], r["genero"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("genero").annotate(total=Count("id")).order_by("-total")
    ]
    por_estado = [
        {"label": ESTADO_MAP.get(r["estado"], r["estado"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("estado").annotate(total=Count("id")).order_by("-total")
    ]
    por_conv = [
        {"label": r["convocatoria__titulo"] or "Sin dato", "total": r["total"]}
        for r in qs.values("convocatoria__titulo").annotate(total=Count("id")).order_by("-total")
    ]

    user_ids = qs.values_list("user_id", flat=True).distinct()
    personas = {
        p.user_id: p
        for p in PersonaHumana.objects.filter(user_id__in=user_ids)
    }

    genero_persona = {}
    rango_etario   = {}
    residencia     = {}

    for p in qs:
        ph = personas.get(p.user_id)

        g_code = getattr(ph, "genero", None) or ""
        g = GENERO_PH_MAP.get(g_code, "Sin dato") if g_code else "Sin dato"
        genero_persona[g] = genero_persona.get(g, 0) + 1

        r = _rango(_edad(ph))
        rango_etario[r] = rango_etario.get(r, 0) + 1

        if ph:
            loc = ph.lugar_residencia
            loc = (ph.otro_lugar_residencia or "Otro") if loc == "otro" else (ph.get_lugar_residencia_display() if loc else "Sin dato")
        else:
            loc = "Sin dato"
        residencia[loc] = residencia.get(loc, 0) + 1

    return {
        "por_tipo":        por_tipo,
        "por_genero_proy": por_genero_proy,
        "por_estado":      por_estado,
        "por_conv":        por_conv,
        "genero_persona":  sorted(genero_persona.items()),
        "rango_etario":    sorted(rango_etario.items()),
        "residencia":      sorted(residencia.items(), key=lambda x: -x[1]),
    }


CAMPOS_IMPACTO = [
    ("honorarios_tecnicos",     "Honorarios técnicos"),
    ("honorarios_elenco",       "Honorarios elenco"),
    ("otros_honorarios",        "Otros honorarios"),
    ("insumos",                 "Insumos"),
    ("servicios_audiovisuales", "Servicios audiovisuales"),
    ("servicios_logistica",     "Servicios / logística"),
]


def _impacto(filtros):
    qs = Rendicion.objects.filter(estado="APROBADO").select_related("postulacion__convocatoria")
    if filtros.get("conv"):
        qs = qs.filter(postulacion__convocatoria_id=filtros["conv"])
    if filtros.get("anio"):
        qs = qs.filter(postulacion__fecha_envio__year=filtros["anio"])

    totales = {campo: 0 for campo, _ in CAMPOS_IMPACTO}
    for r in qs:
        for campo, _ in CAMPOS_IMPACTO:
            totales[campo] += float(getattr(r, campo) or 0)

    total_general = sum(totales.values())
    return {
        "impacto_filas": [
            {"label": label, "monto": totales[campo],
             "pct": round(totales[campo] / total_general * 100, 1) if total_general else 0}
            for campo, label in CAMPOS_IMPACTO
        ],
        "impacto_total": total_general,
        "impacto_count": qs.count(),
    }


@staff_member_required
def dashboard(request):
    convocatorias = Convocatoria.objects.order_by("-fecha_inicio")
    anios = (
        Postulacion.objects.exclude(fecha_envio=None)
        .values_list("fecha_envio__year", flat=True)
        .distinct()
        .order_by("-fecha_envio__year")
    )

    filtros = {
        "conv":           request.GET.get("conv", ""),
        "anio":           request.GET.get("anio", ""),
        "tipo":           request.GET.get("tipo", ""),
        "solo_ganadores": request.GET.get("solo_ganadores", "") == "1",
    }

    qs    = _postulaciones_qs(filtros)
    datos = _agrupar(qs)

    return render(request, "estadisticas/dashboard.html", {
        "filtros":       filtros,
        "convocatorias": convocatorias,
        "anios":         anios,
        "tipos":         Postulacion.TIPO_PROYECTO,
        "total":         qs.count(),
        **datos,
        **_impacto(filtros),
    })


@staff_member_required
def exportar(request):
    filtros = {
        "conv":           request.GET.get("conv", ""),
        "anio":           request.GET.get("anio", ""),
        "tipo":           request.GET.get("tipo", ""),
        "solo_ganadores": request.GET.get("solo_ganadores", "") == "1",
    }
    qs = _postulaciones_qs(filtros)

    user_ids = qs.values_list("user_id", flat=True).distinct()
    personas = {p.user_id: p for p in PersonaHumana.objects.filter(user_id__in=user_ids)}

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Postulaciones"
    _write_headers(ws1, [
        "ID", "Convocatoria", "Proyecto", "Tipo", "Género proyecto",
        "Estado", "Fecha envío",
        "Presentante", "CUIT/CUIL", "Género persona",
        "Rango etario", "Lugar de residencia",
    ])
    for p in qs.order_by("convocatoria", "fecha_envio"):
        ph = personas.get(p.user_id)
        loc = ""
        if ph and ph.lugar_residencia:
            loc = ph.otro_lugar_residencia if ph.lugar_residencia == "otro" else ph.get_lugar_residencia_display()
        ws1.append([
            p.id,
            p.convocatoria.titulo,
            p.nombre_proyecto or "",
            p.get_tipo_proyecto_display() if p.tipo_proyecto else "",
            p.get_genero_display() if p.genero else "",
            p.get_estado_display(),
            p.fecha_envio.strftime("%d/%m/%Y") if p.fecha_envio else "",
            ph.nombre_completo if ph else (p.user.get_full_name() or p.user.email),
            ph.cuil_cuit if ph else "",
            ph.get_genero_display() if ph and ph.genero else "Sin dato",
            _rango(_edad(ph)),
            loc or "Sin dato",
        ])
    _autowidth(ws1)

    ws2 = wb.create_sheet("Resumen")
    datos = _agrupar(qs)
    fila = 1
    fila = _tabla_resumen(ws2, fila, "Por convocatoria",
                          [(r["convocatoria__titulo"], r["total"]) for r in datos["por_conv"]])
    fila = _tabla_resumen(ws2, fila, "Por tipo de proyecto",
                          [(r["tipo_proyecto"] or "Sin dato", r["total"]) for r in datos["por_tipo"]])
    fila = _tabla_resumen(ws2, fila, "Por género del proyecto",
                          [(r["genero"] or "Sin dato", r["total"]) for r in datos["por_genero_proy"]])
    fila = _tabla_resumen(ws2, fila, "Por estado",
                          [(r["estado"], r["total"]) for r in datos["por_estado"]])
    fila = _tabla_resumen(ws2, fila, "Género de presentantes",  datos["genero_persona"])
    fila = _tabla_resumen(ws2, fila, "Rango etario",            datos["rango_etario"])
    fila = _tabla_resumen(ws2, fila, "Lugar de residencia",     datos["residencia"])
    _autowidth(ws2)

    # Hoja 3: impacto económico
    imp = _impacto(filtros)
    if imp["impacto_count"] > 0:
        ws3 = wb.create_sheet("Impacto económico")
        _write_headers(ws3, ["Categoría", "Monto ($)", "%"])
        for fila_imp in imp["impacto_filas"]:
            ws3.append([fila_imp["label"], fila_imp["monto"], fila_imp["pct"] / 100])
        ws3.append(["TOTAL", imp["impacto_total"], 1.0])
        for col in [2, 3]:
            for cell in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=col, max_col=col):
                for c in cell:
                    c.number_format = '#,##0' if col == 2 else '0.0%'
        _autowidth(ws3)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    label = "ganadores" if filtros["solo_ganadores"] else "postulaciones"
    resp["Content-Disposition"] = f'attachment; filename="estadisticas_{label}.xlsx"'
    wb.save(resp)
    return resp


def _write_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A6B")
        cell.alignment = Alignment(horizontal="center")


def _tabla_resumen(ws, fila, titulo, filas):
    c1 = ws.cell(row=fila, column=1, value=titulo)
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor="2E5FA3")
    ws.cell(row=fila, column=2).fill = PatternFill("solid", fgColor="2E5FA3")
    fila += 1
    for label, total in filas:
        ws.cell(row=fila, column=1, value=label)
        ws.cell(row=fila, column=2, value=total)
        fila += 1
    return fila + 1


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
