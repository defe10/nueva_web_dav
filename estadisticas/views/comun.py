"""Helpers compartidos por todos los dashboards de estadísticas.

Convención de unidades (ver estadisticas/INDICADORES.md):
- "postulaciones": una fila por proyecto presentado.
- "personas únicas": una fila por presentante, sin repetir aunque
  haya presentado varios proyectos.
- "contrataciones / ítems": cantidades declaradas en rendiciones;
  pueden repetir personas entre categorías.
"""
from datetime import date

from django.db.models import Count

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


RANGOS_ETARIOS = [
    ("Menos de 18", 0,  17),
    ("18-30",       18, 30),
    ("31-40",       31, 40),
    ("41-50",       41, 50),
    ("51 o más",    51, 999),
]

# Únicos estados del flujo que representan proyectos ganadores:
# "seleccionado" (elegido por el jurado) y "finalizado" (seleccionado
# que además completó su rendición). Si una convocatoria incorpora otro
# estado de ganador, agregarlo acá.
ESTADOS_GANADOR = {"seleccionado", "finalizado"}


def pct(parte, total):
    """Porcentaje redondeado a 1 decimal, o None si no hay base."""
    return round(parte * 100 / total, 1) if total else None


def pesos(valor):
    """Formato abreviado en pesos para mostrar junto a las barras."""
    return "$" + f"{int(valor or 0):,}".replace(",", ".")


def filas_barras(pares, con_pesos=False):
    """Convierte pares (label, valor) en filas para el parcial _barras.html,
    calculando el máximo para el ancho relativo."""
    filas = [
        {"label": label, "total": valor, "display": pesos(valor) if con_pesos else valor}
        for label, valor in pares
    ]
    max_total = max((f["total"] for f in filas), default=0)
    return {"filas": filas, "max_total": max_total}


def edad(persona):
    if not persona or not persona.fecha_nacimiento:
        return None
    hoy = date.today()
    fn = persona.fecha_nacimiento
    return hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))


def rango(edad_valor):
    if edad_valor is None:
        return "Sin dato"
    for label, lo, hi in RANGOS_ETARIOS:
        if lo <= edad_valor <= hi:
            return label
    return "Sin dato"


def conteo(qs, campo, mapa):
    """Agrupa un queryset por un campo y devuelve filas {label, total}."""
    return [
        {"label": mapa.get(r[campo]) or (r[campo] or "Sin dato"), "total": r["total"]}
        for r in qs.values(campo).annotate(total=Count("id")).order_by("-total")
    ]


def normalizar_localidad(texto):
    """Unifica el texto libre de localidad ("san pedro ", "SAN PEDRO",
    "San  Pedro") en un único label, para no contar la misma localidad
    varias veces."""
    return " ".join((texto or "").split()).title()


# ──────────────────────────────────────────────────────────────
# Helpers de exportación a Excel
# ──────────────────────────────────────────────────────────────

def write_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A6B")
        cell.alignment = Alignment(horizontal="center")


def tabla_resumen(ws, fila, titulo, filas):
    c1 = ws.cell(row=fila, column=1, value=titulo)
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor="2E5FA3")
    ws.cell(row=fila, column=2).fill = PatternFill("solid", fgColor="2E5FA3")
    fila += 1
    for label, total in filas:
        ws.cell(row=fila, column=1, value=label)
        ws.cell(row=fila, column=2, value=total)
        fila += 1
    return fila + 1


def autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
