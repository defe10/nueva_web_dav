"""Dashboard y exportación de postulaciones del Plan de Fomento.

Unidades:
- Las tablas por convocatoria / tipo / género de proyecto / estado
  cuentan POSTULACIONES.
- Las tablas demográficas (género, rango etario, residencia) cuentan
  PERSONAS ÚNICAS: cada presentante aparece una sola vez aunque haya
  enviado varias postulaciones.
"""
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse

import openpyxl

from convocatorias.models import Postulacion, Convocatoria, Rendicion
from registro_audiovisual.models import PersonaHumana, GENERO_CHOICES

from .comun import (
    ESTADOS_GANADOR, RANGOS_ETARIOS, edad, rango, normalizar_localidad,
    pct, filas_barras, write_headers, tabla_resumen, autowidth,
)
from .impacto import impacto


# Estados que implican haber superado la revisión administrativa.
ESTADOS_ADMITIDA = {"admitido", "evaluacion_jurado", "seleccionado", "no_seleccionado", "finalizado"}

TIPO_MAP      = {v: l for v, l in Postulacion.TIPO_PROYECTO if v}
GENERO_MAP    = {v: l for v, l in Postulacion.GENERO if v}
ESTADO_MAP    = {v: l for v, l in Postulacion.ESTADOS if v}
GENERO_PH_MAP = dict(GENERO_CHOICES)


def postulaciones_qs(filtros):
    qs = Postulacion.objects.exclude(estado="borrador").select_related(
        "convocatoria", "user"
    )
    if filtros.get("conv"):
        qs = qs.filter(convocatoria_id=filtros["conv"])
    if filtros.get("anio"):
        qs = qs.filter(fecha_envio__year=filtros["anio"])
    if filtros.get("tipo"):
        qs = qs.filter(tipo_proyecto=filtros["tipo"])
    if filtros.get("solo_ganadores"):
        qs = qs.filter(estado__in=ESTADOS_GANADOR)
    return qs


def _residencia_label(ph):
    if not ph or not ph.lugar_residencia:
        return "Sin dato"
    if ph.lugar_residencia == "otro":
        return normalizar_localidad(ph.otro_lugar_residencia) or "Otro (sin especificar)"
    return ph.get_lugar_residencia_display()


def agrupar(qs):
    por_tipo = [
        {"label": TIPO_MAP.get(r["tipo_proyecto"], r["tipo_proyecto"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("tipo_proyecto").annotate(total=Count("id")).order_by("-total")
    ]
    por_genero_proy = [
        {"label": GENERO_MAP.get(r["genero"], r["genero"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("genero").annotate(total=Count("id")).order_by("-total")
    ]
    por_estado = [
        {"label": ESTADO_MAP.get(r["estado"], r["estado"] or "Sin dato"), "total": r["total"]}
        for r in qs.values("estado").annotate(total=Count("id")).order_by("-total")
    ]
    por_conv = [
        {"label": r["convocatoria__titulo"] or "Sin dato", "total": r["total"]}
        for r in qs.values("convocatoria__titulo").annotate(total=Count("id")).order_by("-total")
    ]

    # Demográficos: se cuenta cada presentante UNA vez, aunque tenga
    # varias postulaciones en el resultado filtrado.
    user_ids = set(qs.values_list("user_id", flat=True))
    personas = {
        p.user_id: p
        for p in PersonaHumana.objects.filter(user_id__in=user_ids)
    }

    genero_persona = {}
    rango_etario   = {}
    residencia     = {}

    for uid in user_ids:
        ph = personas.get(uid)

        g_code = getattr(ph, "genero", None) or ""
        g = GENERO_PH_MAP.get(g_code, "Sin dato") if g_code else "Sin dato"
        genero_persona[g] = genero_persona.get(g, 0) + 1

        r = rango(edad(ph))
        rango_etario[r] = rango_etario.get(r, 0) + 1

        loc = _residencia_label(ph)
        residencia[loc] = residencia.get(loc, 0) + 1

    return {
        "por_tipo":        por_tipo,
        "por_genero_proy": por_genero_proy,
        "por_estado":      por_estado,
        "por_conv":        por_conv,
        "presentantes_unicos": len(user_ids),
        "sin_registro":    len(user_ids) - len(personas),
        "genero_persona":  sorted(genero_persona.items()),
        "rango_etario":    [
            (label, rango_etario[label])
            for label in [r[0] for r in RANGOS_ETARIOS] + ["Sin dato"]
            if label in rango_etario
        ],
        "residencia":      sorted(residencia.items(), key=lambda x: -x[1]),
    }


def tasas(qs):
    """Tasas del embudo de postulaciones, sobre el resultado filtrado.

    - admisión:     superaron la revisión administrativa / total
    - selección:    ganadoras / total
    - finalización: finalizadas / ganadoras
    - rendición:    rendiciones aprobadas / ganadoras
    """
    total = qs.count()
    por_estado = {r["estado"]: r["total"] for r in qs.values("estado").annotate(total=Count("id"))}

    admitidas   = sum(por_estado.get(e, 0) for e in ESTADOS_ADMITIDA)
    ganadoras   = sum(por_estado.get(e, 0) for e in ESTADOS_GANADOR)
    finalizadas = por_estado.get("finalizado", 0)
    rendidas_ok = Rendicion.objects.filter(postulacion__in=qs, estado="APROBADO").count()

    return {"tasas": [
        {"nombre": "Admisión",     "valor": pct(admitidas, total),       "detalle": f"{admitidas} de {total} postulaciones"},
        {"nombre": "Selección",    "valor": pct(ganadoras, total),       "detalle": f"{ganadoras} de {total} postulaciones"},
        {"nombre": "Finalización", "valor": pct(finalizadas, ganadoras), "detalle": f"{finalizadas} de {ganadoras} ganadoras"},
        {"nombre": "Rendición aprobada", "valor": pct(rendidas_ok, ganadoras), "detalle": f"{rendidas_ok} de {ganadoras} ganadoras"},
    ]}


def evolucion_anual(filtros):
    """Series anuales para los gráficos de barras. Ignora el filtro de año
    (la evolución compara años entre sí) pero respeta los demás."""
    filtros_sin_anio = {**filtros, "anio": ""}
    qs = postulaciones_qs(filtros_sin_anio).exclude(fecha_envio=None)

    por_anio = (
        qs.values("fecha_envio__year")
        .annotate(total=Count("id"),
                  ganadoras=Count("id", filter=Q(estado__in=ESTADOS_GANADOR)))
        .order_by("fecha_envio__year")
    )
    aprobado_por_anio = (
        Rendicion.objects.filter(estado="APROBADO", fecha_aprobacion__isnull=False)
        .filter(postulacion__in=postulaciones_qs(filtros_sin_anio))
        .values("fecha_aprobacion__year")
        .annotate(total=Sum("honorarios_tecnicos") + Sum("honorarios_elenco")
                        + Sum("otros_honorarios") + Sum("insumos")
                        + Sum("servicios_audiovisuales") + Sum("servicios_logistica"))
        .order_by("fecha_aprobacion__year")
    )

    return {
        "evolucion_postulaciones": filas_barras(
            [(r["fecha_envio__year"], r["total"]) for r in por_anio]),
        "evolucion_ganadoras": filas_barras(
            [(r["fecha_envio__year"], r["ganadoras"]) for r in por_anio]),
        "evolucion_monto_aprobado": filas_barras(
            [(r["fecha_aprobacion__year"], r["total"] or 0) for r in aprobado_por_anio],
            con_pesos=True),
    }


@staff_member_required
def dashboard(request):
    convocatorias = Convocatoria.objects.order_by("-fecha_inicio")
    anios = (
        Postulacion.objects.exclude(fecha_envio=None)
        .values_list("fecha_envio__year", flat=True)
        .distinct()
        .order_by("-fecha_envio__year")
    )

    filtros = {
        "conv":           request.GET.get("conv", ""),
        "anio":           request.GET.get("anio", ""),
        "tipo":           request.GET.get("tipo", ""),
        "solo_ganadores": request.GET.get("solo_ganadores", "") == "1",
    }

    qs    = postulaciones_qs(filtros)
    datos = agrupar(qs)

    return render(request, "estadisticas/dashboard.html", {
        "filtros":       filtros,
        "convocatorias": convocatorias,
        "anios":         anios,
        "tipos":         Postulacion.TIPO_PROYECTO,
        "total":         qs.count(),
        **datos,
        **tasas(qs),
        **evolucion_anual(filtros),
        **impacto(filtros),
    })


@staff_member_required
def exportar(request):
    filtros = {
        "conv":           request.GET.get("conv", ""),
        "anio":           request.GET.get("anio", ""),
        "tipo":           request.GET.get("tipo", ""),
        "solo_ganadores": request.GET.get("solo_ganadores", "") == "1",
    }
    qs = postulaciones_qs(filtros)

    user_ids = qs.values_list("user_id", flat=True).distinct()
    personas = {p.user_id: p for p in PersonaHumana.objects.filter(user_id__in=user_ids)}

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Postulaciones"
    write_headers(ws1, [
        "ID", "Convocatoria", "Proyecto", "Tipo", "Género proyecto",
        "Estado", "Fecha envío",
        "Presentante", "CUIT/CUIL", "Género persona",
        "Rango etario", "Lugar de residencia",
    ])
    for p in qs.order_by("convocatoria", "fecha_envio"):
        ph = personas.get(p.user_id)
        loc = _residencia_label(ph)
        ws1.append([
            p.id,
            p.convocatoria.titulo,
            p.nombre_proyecto or "",
            p.get_tipo_proyecto_display() if p.tipo_proyecto else "",
            p.get_genero_display() if p.genero else "",
            p.get_estado_display(),
            p.fecha_envio.strftime("%d/%m/%Y") if p.fecha_envio else "",
            ph.nombre_completo if ph else (p.user.get_full_name() or p.user.email),
            ph.cuil_cuit if ph else "",
            ph.get_genero_display() if ph and ph.genero else "Sin dato",
            rango(edad(ph)),
            loc,
        ])
    autowidth(ws1)

    ws2 = wb.create_sheet("Resumen")
    datos = agrupar(qs)
    fila = 1
    fila = tabla_resumen(ws2, fila, "Por convocatoria (postulaciones)",
                         [(r["label"], r["total"]) for r in datos["por_conv"]])
    fila = tabla_resumen(ws2, fila, "Por tipo de proyecto (postulaciones)",
                         [(r["label"], r["total"]) for r in datos["por_tipo"]])
    fila = tabla_resumen(ws2, fila, "Por género del proyecto (postulaciones)",
                         [(r["label"], r["total"]) for r in datos["por_genero_proy"]])
    fila = tabla_resumen(ws2, fila, "Por estado (postulaciones)",
                         [(r["label"], r["total"]) for r in datos["por_estado"]])
    fila = tabla_resumen(ws2, fila, "Género (presentantes únicos)",  datos["genero_persona"])
    fila = tabla_resumen(ws2, fila, "Rango etario (presentantes únicos)", datos["rango_etario"])
    fila = tabla_resumen(ws2, fila, "Lugar de residencia (presentantes únicos)", datos["residencia"])
    autowidth(ws2)

    # Hoja 3: impacto económico
    imp = impacto(filtros)
    if imp["impacto_count"] > 0:
        ws3 = wb.create_sheet("Impacto económico")
        write_headers(ws3, ["Categoría", "Monto ($)", "Contrataciones / ítems", "%"])
        for fila_imp in imp["impacto_filas"]:
            ws3.append([fila_imp["label"], fila_imp["monto"], fila_imp["cantidad"], fila_imp["pct"] / 100])
        ws3.append(["TOTAL", imp["impacto_total"], imp["impacto_total_cantidad"], 1.0])
        for col in [2, 3, 4]:
            for cell in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=col, max_col=col):
                for c in cell:
                    c.number_format = '#,##0' if col in (2, 3) else '0.0%'
        autowidth(ws3)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    label = "ganadores" if filtros["solo_ganadores"] else "postulaciones"
    resp["Content-Disposition"] = f'attachment; filename="estadisticas_{label}.xlsx"'
    wb.save(resp)
    return resp
