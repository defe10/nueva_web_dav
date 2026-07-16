"""Dashboard y exportación del Registro Audiovisual.

Unidades: todas las tablas cuentan PERSONAS (humanas o jurídicas)
inscriptas en el registro, salvo "técnicos disponibles por área", donde
una persona puede aparecer en dos áreas (principal y secundaria).

Filtros: localidad (residencia de PH / localidad fiscal de PJ) y
área de desempeño (principal o secundaria, solo personas humanas).
"""
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Q
from django.http import HttpResponse

import openpyxl

from registro_audiovisual.models import (
    PersonaHumana, PersonaJuridica, GENERO_CHOICES,
    AREA_DESEMPENO, AREA_DESEMPENO_PPJJ, LUGARES_RESIDENCIA,
    NIVEL_EDUCATIVO_CHOICES, SITUACION_IVA, TIPO_PERSONA_JURIDICA_CHOICES,
)

from .comun import (
    conteo, normalizar_localidad, filas_barras,
    write_headers, tabla_resumen, autowidth, edad, rango,
)


AREA_MAP      = dict(AREA_DESEMPENO)
AREA_PPJJ_MAP = dict(AREA_DESEMPENO_PPJJ)
LUGAR_MAP     = dict(LUGARES_RESIDENCIA)
NIVEL_MAP     = dict(NIVEL_EDUCATIVO_CHOICES)
IVA_MAP       = dict(SITUACION_IVA)
TIPO_PJ_MAP   = dict(TIPO_PERSONA_JURIDICA_CHOICES)
GENERO_PH_MAP = dict(GENERO_CHOICES)


def _filtros(request):
    return {
        "loc":  request.GET.get("loc", ""),
        "area": request.GET.get("area", ""),
    }


def _querysets(filtros):
    ph = PersonaHumana.objects.all()
    pj = PersonaJuridica.objects.all()
    if filtros["loc"]:
        ph = ph.filter(lugar_residencia=filtros["loc"])
        pj = pj.filter(localidad_fiscal=filtros["loc"])
    if filtros["area"]:
        ph = ph.filter(Q(area_desempeno_1=filtros["area"]) | Q(area_desempeno_2=filtros["area"]))
    return ph, pj


def _residencia(ph_qs):
    """Ubicación territorial desagregando "Otro" por el texto libre
    cargado (normalizado), en vez de agrupar localidades distintas
    en una sola fila."""
    filas = {}
    for lugar, otro in ph_qs.values_list("lugar_residencia", "otro_lugar_residencia"):
        if lugar == "otro":
            label = normalizar_localidad(otro) or "Otro (sin especificar)"
        else:
            label = LUGAR_MAP.get(lugar) or (lugar or "Sin dato")
        filas[label] = filas.get(label, 0) + 1
    return [
        {"label": label, "total": total}
        for label, total in sorted(filas.items(), key=lambda x: -x[1])
    ]


def _tecnicos_por_area(ph_qs):
    """Personas disponibles por área, contando área principal y secundaria."""
    filas = {}
    for campo in ("area_desempeno_1", "area_desempeno_2"):
        agrupado = (
            ph_qs.exclude(**{f"{campo}__isnull": True})
            .exclude(**{campo: ""})
            .values(campo)
            .annotate(total=Count("id"))
        )
        for r in agrupado:
            label = AREA_MAP.get(r[campo], r[campo])
            filas[label] = filas.get(label, 0) + r["total"]
    return [
        {"label": label, "total": total}
        for label, total in sorted(filas.items(), key=lambda x: -x[1])
    ]


def _altas_por_anio(ph_qs, pj_qs):
    anios = {}
    for qs in (ph_qs, pj_qs):
        for r in qs.values("fecha_creacion__year").annotate(total=Count("id")):
            anio = r["fecha_creacion__year"]
            anios[anio] = anios.get(anio, 0) + r["total"]
    return filas_barras(sorted(anios.items()))


@staff_member_required
def dashboard_registro(request):
    filtros = _filtros(request)
    ph, pj = _querysets(filtros)

    total_ph = ph.count()
    total_pj = pj.count()
    localidades = (
        ph.exclude(lugar_residencia="").values("lugar_residencia").distinct().count()
    )

    return render(request, "estadisticas/dashboard_registro.html", {
        "filtros":      filtros,
        "lugares":      LUGARES_RESIDENCIA,
        "areas":        AREA_DESEMPENO,
        "total_ph":     total_ph,
        "total_pj":     total_pj,
        "localidades":  localidades,
        "evolucion_altas": _altas_por_anio(ph, pj),
        # Personas humanas
        "roles":            conteo(ph, "area_desempeno_1", AREA_MAP),
        "tecnicos":         _tecnicos_por_area(ph),
        "residencia_ph":    _residencia(ph),
        "genero_ph":        conteo(ph, "genero", GENERO_PH_MAP),
        "nivel_educativo":  conteo(ph, "nivel_educativo", NIVEL_MAP),
        "situacion_iva_ph": conteo(ph, "situacion_iva", IVA_MAP),
        # Personas jurídicas
        "tipo_pj":      conteo(pj, "tipo_persona_juridica", TIPO_PJ_MAP),
        "area_pj":      conteo(pj, "area_desempeno_JJPP_1", AREA_PPJJ_MAP),
        "localidad_pj": conteo(pj, "localidad_fiscal", LUGAR_MAP),
    })


@staff_member_required
def exportar_registro(request):
    filtros = _filtros(request)
    ph, pj = _querysets(filtros)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Personas humanas"
    write_headers(ws1, [
        "Nombre", "Apellido", "CUIL/CUIT", "Género", "Rango etario",
        "Nivel educativo", "Localidad", "Área principal", "Área secundaria",
        "Área cultural", "Situación IVA", "Email", "Teléfono",
    ])
    for p in ph.order_by("apellido", "nombre"):
        if p.lugar_residencia == "otro":
            loc = normalizar_localidad(p.otro_lugar_residencia) or "Otro (sin especificar)"
        else:
            loc = p.get_lugar_residencia_display() if p.lugar_residencia else "Sin dato"
        ws1.append([
            p.nombre, p.apellido, p.cuil_cuit,
            p.get_genero_display() if p.genero else "Sin dato",
            rango(edad(p)),
            p.get_nivel_educativo_display() if p.nivel_educativo else "Sin dato",
            loc,
            p.get_area_desempeno_1_display() if p.area_desempeno_1 else "",
            p.get_area_desempeno_2_display() if p.area_desempeno_2 else "",
            p.get_area_cultural_display() if p.area_cultural else "",
            p.get_situacion_iva_display() if p.situacion_iva else "",
            p.email, p.telefono,
        ])
    autowidth(ws1)

    ws2 = wb.create_sheet("Personas jurídicas")
    write_headers(ws2, [
        "Razón social", "Nombre comercial", "CUIT", "Tipo", "Área principal",
        "Localidad fiscal", "Situación IVA", "Fecha constitución", "Email", "Teléfono",
    ])
    for p in pj.order_by("razon_social"):
        ws2.append([
            p.razon_social, p.nombre_comercial or "", p.cuil_cuit,
            p.get_tipo_persona_juridica_display(),
            p.get_area_desempeno_JJPP_1_display() if p.area_desempeno_JJPP_1 else "",
            p.get_localidad_fiscal_display() if p.localidad_fiscal else "Sin dato",
            p.get_situacion_iva_display() if p.situacion_iva else "",
            p.fecha_constitucion.strftime("%d/%m/%Y") if p.fecha_constitucion else "",
            p.email, p.telefono,
        ])
    autowidth(ws2)

    ws3 = wb.create_sheet("Resumen")
    fila = 1
    fila = tabla_resumen(ws3, fila, "Roles predominantes (área principal)",
                         [(r["label"], r["total"]) for r in conteo(ph, "area_desempeno_1", AREA_MAP)])
    fila = tabla_resumen(ws3, fila, "Técnicos disponibles por área",
                         [(r["label"], r["total"]) for r in _tecnicos_por_area(ph)])
    fila = tabla_resumen(ws3, fila, "Ubicación territorial",
                         [(r["label"], r["total"]) for r in _residencia(ph)])
    fila = tabla_resumen(ws3, fila, "Género",
                         [(r["label"], r["total"]) for r in conteo(ph, "genero", GENERO_PH_MAP)])
    fila = tabla_resumen(ws3, fila, "Nivel educativo",
                         [(r["label"], r["total"]) for r in conteo(ph, "nivel_educativo", NIVEL_MAP)])
    fila = tabla_resumen(ws3, fila, "Tipo de persona jurídica",
                         [(r["label"], r["total"]) for r in conteo(pj, "tipo_persona_juridica", TIPO_PJ_MAP)])
    autowidth(ws3)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="estadisticas_registro.xlsx"'
    wb.save(resp)
    return resp
