# exencion/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone

from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.urls import reverse
from convocatorias.models import Convocatoria
from registro_audiovisual.models import PersonaHumana, PersonaJuridica

from .models import Exencion, ExencionDocumento, ObservacionAdministrativaExencion, PadronPublicoExencion
from .utils import datos_fiscales_completos


# ============================================================
# CONFIG
# ============================================================
MAX_DOCS_SUBSANACION = 10  # límite solo para documentos de subsanación

TIPOS_DOC_EXENCION = ["CV", "DNI", "CONSTANCIA_ARCA"]


# ============================================================
# HELPERS
# ============================================================
def _es_pdf(archivo):
    nombre = (archivo.name or "").lower()
    content_type = (getattr(archivo, "content_type", "") or "").lower()
    return nombre.endswith(".pdf") or ("pdf" in content_type)


def _cupos_restantes_subsanacion(exencion):
    total_subs = ExencionDocumento.objects.filter(exencion=exencion, es_subsanacion=True).count()
    return MAX_DOCS_SUBSANACION - total_subs


# ============================================================
# PASO 1 — INICIAR SOLICITUD
# ============================================================
def iniciar_solicitud(request, convocatoria_id=None):

    if not request.user.is_authenticated:
        messages.warning(request, "Para solicitar una exención impositiva necesitás ingresar con tu usuario.")
        return redirect(f"/usuarios/login/?next={request.path}")

    convocatoria = None
    if convocatoria_id is not None:
        convocatoria = get_object_or_404(Convocatoria, id=convocatoria_id)

    user = request.user

    persona_humana = PersonaHumana.objects.filter(user=user).first()
    persona_juridica = PersonaJuridica.objects.filter(user=user).first()

    if not (persona_humana or persona_juridica):
        messages.warning(request, "Para solicitar una exención impositiva primero debés completar tu Registro Audiovisual.")
        next_url = request.path + "?confirmed=1"
        return redirect(reverse("registro_audiovisual:seleccionar_tipo_registro") + f"?next={next_url}")

    persona = persona_humana or persona_juridica

    if not datos_fiscales_completos(persona):
        messages.warning(
            request,
            "Para solicitar la exención impositiva es necesario completar previamente "
            "todos los datos fiscales en el Registro Audiovisual "
            "(Situación IVA, Actividad DGR, Domicilio fiscal, Localidad fiscal y Código Postal). "
            'Si elegiste "Ninguna/Ninguno" en IVA o DGR, se considera incompleto.'
        )
        edit_view = (
            "registro_audiovisual:editar_persona_humana"
            if persona_humana
            else "registro_audiovisual:editar_persona_juridica"
        )
        next_url = request.path + "?confirmed=1"
        tab = "fiscales" if persona_humana else "generales"
        return redirect(reverse(edit_view) + f"?next={next_url}&tab={tab}")

    # Revisar si ya existe una exención para este usuario/convocatoria
    from django.utils import timezone as tz
    hoy = tz.now().date()

    exencion_existente = (
        Exencion.objects
        .filter(user=user, convocatoria=convocatoria)
        .order_by("-fecha_creacion")
        .first()
    )

    if exencion_existente:
        estado = exencion_existente.estado
        if estado == "BORRADOR":
            return redirect("exencion:documentacion", exencion_id=exencion_existente.id)
        if estado in ("ENVIADA", "OBSERVADA"):
            messages.info(request, "Tu solicitud ya fue enviada y está en revisión administrativa.")
            return redirect("exencion:completada", exencion_id=exencion_existente.id)
        if estado == "APROBADA":
            vencimiento = exencion_existente.fecha_vencimiento
            if vencimiento and vencimiento > hoy:
                messages.success(
                    request,
                    "Ya tenés una constancia de exención vigente. "
                    "Podés descargarla desde tu panel de usuario."
                )
                return redirect("usuarios:panel_usuario")
            # Vencida → permitir renovación (crear nueva abajo)

    # Usuario con registro completo: ofrecer revisión de datos (solo si no viene de confirmar)
    if not request.GET.get("confirmed"):
        next_url = request.path + "?confirmed=1"
        return redirect(reverse("registro_audiovisual:confirmar_datos") + f"?next={next_url}")

    # Crear nueva exencion (renovación o primera vez)
    exencion = Exencion(
        user=user,
        convocatoria=convocatoria,
        persona_humana=persona_humana,
        persona_juridica=persona_juridica,
        nombre_razon_social=(
            getattr(persona, "nombre_completo", None)
            or getattr(persona, "razon_social", "")
            or ""
        ),
        email=getattr(persona, "email", "") or user.email or "",
        cuit=getattr(persona, "cuil_cuit", "") or "",
        domicilio_fiscal=getattr(persona, "domicilio_fiscal", "") or "",
        localidad_fiscal=getattr(persona, "localidad_fiscal", "") or "",
        codigo_postal_fiscal=getattr(persona, "codigo_postal_fiscal", "") or "",
        actividad_dgr=getattr(persona, "actividad_dgr", "") or "",
        estado="BORRADOR",
    )
    exencion.save()

    return redirect("exencion:documentacion", exencion_id=exencion.id)


# ============================================================
# DOCUMENTACIÓN INICIAL (pantalla)
# ============================================================
@login_required(login_url="/usuarios/login/")
def subir_documentacion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("convocatorias:convocatorias_home")

    persona = exencion.persona_humana or exencion.persona_juridica

    if not datos_fiscales_completos(persona):
        messages.warning(
            request,
            "Para continuar con la solicitud de exención debés completar previamente "
            "todos los datos fiscales en tu Registro Audiovisual. "
            'Si elegiste "Ninguna/Ninguno" en IVA o DGR, se considera incompleto.'
        )
        next_url = request.path
        if exencion.persona_humana:
            return redirect(reverse("registro_audiovisual:editar_persona_humana") + f"?next={next_url}&tab=fiscales")
        return redirect(reverse("registro_audiovisual:editar_persona_juridica") + f"?next={next_url}&tab=generales")

    documentos_pendientes = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=False,
        estado="PENDIENTE",
    ).order_by("-fecha_subida")

    documentos_enviados = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=False,
        estado="ENVIADO",
    ).order_by("-fecha_subida")

    restantes = _cupos_restantes_subsanacion(exencion)

    docs_iniciales = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=False,
    )
    doc_cv   = docs_iniciales.filter(tipo="CV").first()
    doc_dni  = docs_iniciales.filter(tipo="DNI").first()
    doc_arca = docs_iniciales.filter(tipo="CONSTANCIA_ARCA").first()

    todos_subidos = all([doc_cv, doc_dni, doc_arca])
    alguno_enviado = docs_iniciales.filter(estado="ENVIADO").exists()

    return render(
        request,
        "exencion/documentacion.html",
        {
            "exencion": exencion,
            "convocatoria": exencion.convocatoria,
            "documentos_pendientes": documentos_pendientes,
            "documentos_enviados": documentos_enviados,
            "doc_cv": doc_cv,
            "doc_dni": doc_dni,
            "doc_arca": doc_arca,
            "todos_subidos": todos_subidos,
            "alguno_enviado": alguno_enviado,
            "max_archivos": MAX_DOCS_SUBSANACION,
            "restantes": restantes,
        }
    )


# ============================================================
# AGREGAR DOCUMENTACIÓN INICIAL (POST)
# ============================================================
@login_required(login_url="/usuarios/login/")
def agregar_documentacion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("convocatorias:convocatorias_home")

    if request.method != "POST":
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    tipo = request.POST.get("tipo")
    archivo = request.FILES.get("archivo")

    if not tipo or tipo not in TIPOS_DOC_EXENCION:
        messages.error(request, "Tipo de documento inválido.")
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    if not archivo:
        messages.error(request, "Seleccioná un archivo para subir.")
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    if not _es_pdf(archivo):
        messages.error(request, f"El archivo '{archivo.name}' no es PDF.")
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    # Si ya existe uno pendiente del mismo tipo, reemplazarlo; si está enviado, no permitir
    existente = ExencionDocumento.objects.filter(
        exencion=exencion,
        tipo=tipo,
        es_subsanacion=False,
    ).first()

    if existente and existente.estado == "ENVIADO":
        messages.error(request, "No podés reemplazar un documento ya enviado.")
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    if existente:
        existente.archivo = archivo
        try:
            existente.full_clean()
            existente.save()
        except Exception as e:
            messages.error(request, str(e))
    else:
        doc = ExencionDocumento(
            exencion=exencion,
            tipo=tipo,
            archivo=archivo,
            es_subsanacion=False,
            estado="PENDIENTE",
        )
        try:
            doc.full_clean()
            doc.save()
        except Exception as e:
            messages.error(request, str(e))

    return redirect("exencion:documentacion", exencion_id=exencion.id)


# ============================================================
# CONFIRMAR ENVÍO DOCUMENTACIÓN INICIAL
# ============================================================
@login_required(login_url="/usuarios/login/")
def confirmar_documentacion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("convocatorias:convocatorias_home")

    if request.method != "POST":
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    # Exigir los 3 tipos de documento
    docs_iniciales = ExencionDocumento.objects.filter(exencion=exencion, es_subsanacion=False)
    tipos_subidos = set(docs_iniciales.values_list("tipo", flat=True))
    faltantes = [t for t in TIPOS_DOC_EXENCION if t not in tipos_subidos]
    if faltantes:
        messages.error(request, "Debés subir los tres documentos requeridos antes de enviar.")
        return redirect("exencion:documentacion", exencion_id=exencion.id)

    qs_pendientes = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=False,
        estado="PENDIENTE",
    )

    # Si no hay pendientes pero ya hay enviados, no vuelvas a "enviar": redirigí a completada
    if not qs_pendientes.exists():
        messages.info(request, "Tu documentación ya fue enviada.")
        return redirect("exencion:completada", exencion_id=exencion.id)

    ahora = timezone.now()
    qs_pendientes.update(estado="ENVIADO", fecha_envio=ahora)

    exencion.estado = "ENVIADA"
    exencion.save(update_fields=["estado"])

    return redirect("exencion:completada", exencion_id=exencion.id)


# ============================================================
# ELIMINAR DOCUMENTO (inicial o subsanado)
# ============================================================
@login_required(login_url="/usuarios/login/")
def eliminar_documento(request, documento_id):

    doc = get_object_or_404(ExencionDocumento, id=documento_id)

    if doc.exencion.user != request.user:
        return redirect("usuarios:panel_usuario")

    if doc.estado != "PENDIENTE":
        messages.error(request, "No se puede eliminar un documento ya enviado.")
        return redirect("usuarios:panel_usuario")

    exencion_id = doc.exencion_id
    es_sub = doc.es_subsanacion

    doc.delete()
    messages.success(request, "Documento eliminado.")

    if es_sub:
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion_id)
    return redirect("exencion:documentacion", exencion_id=exencion_id)


# ============================================================
# CONFIRMACIÓN FINAL (pantalla)
# ============================================================
@login_required(login_url="/usuarios/login/")
def solicitud_completada(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("convocatorias:convocatorias_home")

    return render(
        request,
        "exencion/completada.html",
        {"exencion": exencion, "convocatoria": exencion.convocatoria}
    )


# ============================================================
# SUBSANACIÓN (pantalla)
# ============================================================
@login_required(login_url="/usuarios/login/")
def subir_documento_subsanado_exencion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)

    if exencion.user != request.user:
        return redirect("usuarios:panel_usuario")

    observaciones_pendientes = list(
        ObservacionAdministrativaExencion.objects
        .filter(exencion=exencion, subsanada=False)
        .order_by("-fecha_creacion")
    )

    documentos_pendientes = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=True,
        estado="PENDIENTE",
    ).order_by("-fecha_subida")

    documentos_enviados = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=True,
        estado="ENVIADO",
    ).order_by("-fecha_subida")

    restantes = _cupos_restantes_subsanacion(exencion)

    return render(
        request,
        "exencion/subir_documento_subsanado_exencion.html",
        {
            "exencion": exencion,
            "observaciones_pendientes": observaciones_pendientes,
            "documentos_pendientes": documentos_pendientes,
            "documentos_enviados": documentos_enviados,
            "max_archivos": MAX_DOCS_SUBSANACION,
            "restantes": restantes,
        }
    )


# ============================================================
# AGREGAR DOCUMENTOS SUBSANADOS (POST)
# ============================================================
@login_required(login_url="/usuarios/login/")
def agregar_documento_subsanado_exencion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("usuarios:panel_usuario")

    if request.method != "POST":
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    archivos = request.FILES.getlist("archivos")
    if not archivos:
        messages.error(request, "Seleccioná al menos un archivo para subir.")
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    restantes = _cupos_restantes_subsanacion(exencion)

    if restantes <= 0:
        messages.error(request, f"Ya alcanzaste el máximo de {MAX_DOCS_SUBSANACION} archivos de subsanación para esta exención.")
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    if len(archivos) > restantes:
        messages.error(
            request,
            f"Podés subir como máximo {restantes} archivo(s) más de subsanación. Límite: {MAX_DOCS_SUBSANACION}."
        )
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    guardados = 0
    for archivo in archivos:
        if not _es_pdf(archivo):
            messages.error(request, f"El archivo '{archivo.name}' no es PDF.")
            continue

        doc = ExencionDocumento(
            exencion=exencion,
            archivo=archivo,
            es_subsanacion=True,
            estado="PENDIENTE",
        )
        try:
            doc.full_clean()
            doc.save()
            guardados += 1
        except Exception as e:
            messages.error(request, str(e))

    if guardados > 0:
        messages.success(request, f"Se agregaron {guardados} archivo(s).")
    else:
        messages.error(request, "No se pudo subir ningún archivo. Revisá los errores.")

    return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)


# ============================================================
# CONFIRMAR ENVÍO SUBSANACIÓN
# ============================================================
@login_required(login_url="/usuarios/login/")
def confirmar_documento_subsanado_exencion(request, exencion_id):

    exencion = get_object_or_404(Exencion, id=exencion_id)
    if exencion.user != request.user:
        return redirect("usuarios:panel_usuario")

    if request.method != "POST":
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    # ✅ NUEVO: bloquear si no hay observaciones pendientes
    if not ObservacionAdministrativaExencion.objects.filter(
        exencion=exencion,
        subsanada=False
    ).exists():
        messages.error(request, "No tenés observaciones pendientes para subsanar.")
        return redirect("usuarios:panel_usuario")

    # ✅ NUEVO: exigir al menos 1 documento de subsanación cargado
    if not ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=True,
    ).exists():
        messages.error(request, "Debés subir al menos un archivo PDF de subsanación antes de enviar.")
        return redirect("exencion:subir_documento_subsanado_exencion", exencion_id=exencion.id)

    qs_pendientes = ExencionDocumento.objects.filter(
        exencion=exencion,
        es_subsanacion=True,
        estado="PENDIENTE",
    )

    if not qs_pendientes.exists():
        messages.info(request, "Tu subsanación ya fue enviada.")
        return redirect("usuarios:panel_usuario")

    ahora = timezone.now()
    qs_pendientes.update(estado="ENVIADO", fecha_envio=ahora)

    ObservacionAdministrativaExencion.objects.filter(
        exencion=exencion, subsanada=False
    ).update(subsanada=True)

    exencion.estado = "ENVIADA"
    exencion.save(update_fields=["estado"])

    messages.success(request, "Subsanación enviada correctamente.")
    return redirect("usuarios:panel_usuario")


# ============================================================
# PADRÓN PÚBLICO (QR) + EXCEL (interno)
# ============================================================

def padron_publico_exenciones(request):
    """
    Padrón público (para QR): SOLO aprobadas.
    Sin links, sin admin, sin datos sensibles.
    """
    q = (request.GET.get("q") or "").strip()

    qs = Exencion.objects.filter(
        estado="APROBADA"
    ).order_by("-fecha_emision", "-fecha_creacion")

    if q:
        qs = qs.filter(
            Q(cuit__icontains=q) |
            Q(nombre_razon_social__icontains=q)
        )

    rows = []
    for ex in qs:
        rows.append({
            "id": ex.id,
            "nombre_razon_social": ex.nombre_razon_social or "",
            "cuit": ex.cuit or "",
            "fecha_emision": ex.fecha_emision,
            "fecha_vencimiento": ex.fecha_vencimiento,
            "estado": ex.estado or "",
        })

    return render(request, "exencion/padron_publico_exenciones.html", {
        "q": q,
        "rows": rows,
    })


@staff_member_required
def padron_exenciones_excel(request):
    """
    Excel COMPLETO (interno) con TODOS los campos del modelo Exencion.
    Exporta SOLO aprobadas.
    """
    q = (request.GET.get("q") or "").strip()

    qs = Exencion.objects.filter(
        estado="APROBADA"
    ).order_by("-fecha_emision", "-fecha_creacion")

    if q:
        qs = qs.filter(
            Q(cuit__icontains=q) |
            Q(nombre_razon_social__icontains=q)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Exenciones Aprobadas"

    def fmt_value(v):
        if v is None:
            return ""
        try:
            return timezone.localtime(v).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        try:
            return v.strftime("%d/%m/%Y")
        except Exception:
            pass
        try:
            if hasattr(v, "name"):  # FileField
                return v.name or ""
        except Exception:
            pass
        return str(v)

    field_names = [getattr(f, "attname", f.name) for f in Exencion._meta.fields]
    ws.append(field_names)

    for obj in qs:
        ws.append([fmt_value(getattr(obj, n, "")) for n in field_names])

    for col_idx, h in enumerate(field_names, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(h)) + 2, 14), 45)

    filename = "padron_exenciones_aprobadas_completo.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def ir_a_padron_publico_exenciones(request):
    return redirect("exencion:padron_publico_exenciones")
