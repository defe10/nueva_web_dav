from django.contrib import admin, messages
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.conf import settings

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    ConvocatoriaFormacion,
    MiembroFormador,
    ConfiguracionInscripcionFormacion,
    InscripcionFormacion,
    ObservacionFormacion,
)


# ==========================================
# INLINE: miembros dentro de la convocatoria
# ==========================================

class MiembroFormadorInline(admin.TabularInline):
    model = MiembroFormador
    extra = 0
    fields = ["nombre", "orden", "foto", "bio"]


class ConfiguracionInscripcionInline(admin.StackedInline):
    model = ConfiguracionInscripcionFormacion
    extra = 0
    can_delete = False


# ==========================================
# CONVOCATORIA DE FORMACIÓN
# ==========================================

@admin.register(ConvocatoriaFormacion)
class ConvocatoriaFormacionAdmin(admin.ModelAdmin):
    list_display  = ("titulo", "tipo_formacion", "fecha_inicio", "fecha_fin", "cupo_maximo", "orden")
    list_filter   = ("tipo_formacion",)
    search_fields = ("titulo",)
    prepopulated_fields = {"slug": ("titulo",)}
    inlines = [MiembroFormadorInline, ConfiguracionInscripcionInline]


# ==========================================
# INSCRIPCIONES
# ==========================================

class ObservacionFormacionInline(admin.TabularInline):
    model = ObservacionFormacion
    extra = 0
    fields = ["descripcion", "subsanada", "creada_por", "fecha_creacion"]
    readonly_fields = ["fecha_creacion"]


@admin.register(InscripcionFormacion)
class InscripcionFormacionAdmin(admin.ModelAdmin):
    list_display   = ("usuario", "convocatoria", "estado", "contacto_email", "contacto_telefono", "vinculo_sector", "fecha")
    list_filter    = ("estado", "vinculo_sector", "convocatoria")
    search_fields  = ("user__username", "user__email", "nombre", "apellido", "dni", "email", "telefono")
    ordering       = ("-fecha",)
    readonly_fields = ("user", "convocatoria", "fecha")
    actions        = ["marcar_admitido", "marcar_no_admitido", "marcar_lista_espera", "exportar_excel"]
    inlines        = [ObservacionFormacionInline]
    list_select_related = ("user", "convocatoria", "persona_humana", "persona_juridica")
    fieldsets = (
        ("Sistema",              {"fields": ("user", "convocatoria", "estado", "fecha")}),
        ("Registro Audiovisual", {"fields": ("persona_humana", "persona_juridica")}),
        ("Contacto",             {"fields": ("nombre", "apellido", "dni", "email", "telefono", "localidad", "otra_localidad")}),
        ("Perfil",               {"fields": ("vinculo_sector", "declaracion_jurada")}),
    )

    def usuario(self, obj):
        return obj.user.username
    usuario.short_description = "Usuario"

    def contacto_email(self, obj):
        return obj.user.email or obj.email or "—"
    contacto_email.short_description = "Email"

    def contacto_telefono(self, obj):
        if obj.persona_humana_id and getattr(obj.persona_humana, "telefono", None):
            return obj.persona_humana.telefono
        if obj.persona_juridica_id and getattr(obj.persona_juridica, "telefono", None):
            return obj.persona_juridica.telefono
        return obj.telefono or "—"
    contacto_telefono.short_description = "Teléfono"

    # --------------------------------------------------
    # HELPERS
    # --------------------------------------------------
    def _enviar_email_cambio_estado(self, request, inscripcion, nuevo_estado):
        user = inscripcion.user
        destinatario = (user.email or inscripcion.email or "").strip()
        if not destinatario:
            return
        panel_url = request.build_absolute_uri(reverse("usuarios:panel_usuario"))
        asunto = f"Actualización de tu inscripción: {inscripcion.convocatoria.titulo}"
        labels = {
            "admitido":    "¡Tu inscripción fue admitida!",
            "no_admitido": "Tu inscripción no fue admitida.",
            "lista_espera": "Tu inscripción está en lista de espera.",
        }
        texto = (
            f"{labels.get(nuevo_estado, 'Estado actualizado.')}\n\n"
            f"Formación: {inscripcion.convocatoria.titulo}\n"
            f"Panel: {panel_url}"
        )
        try:
            html = render_to_string(
                "formacion/email_cambio_estado.html",
                {
                    "inscripcion": inscripcion,
                    "user": user,
                    "nuevo_estado": nuevo_estado,
                    "panel_url": panel_url,
                    "anio": timezone.now().year,
                },
            )
            email = EmailMultiAlternatives(
                subject=asunto, body=texto,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                to=[destinatario],
            )
            email.attach_alternative(html, "text/html")
            email.send(fail_silently=True)
        except Exception as e:
            messages.warning(request, f"No se pudo enviar email a {destinatario}: {e}")

    def _cambiar_estado(self, request, queryset, nuevo_estado):
        actualizadas = 0
        for ins in queryset:
            if ins.estado != nuevo_estado:
                ins.estado = nuevo_estado
                ins.save(update_fields=["estado"])
                self._enviar_email_cambio_estado(request, ins, nuevo_estado)
                actualizadas += 1
        if actualizadas:
            self.message_user(request, f"{actualizadas} inscripción/es actualizadas a '{nuevo_estado}'.")

    # --------------------------------------------------
    # ACCIONES
    # --------------------------------------------------
    @admin.action(description="✅ Marcar como ADMITIDO")
    def marcar_admitido(self, request, queryset):
        self._cambiar_estado(request, queryset, "admitido")

    @admin.action(description="❌ Marcar como NO ADMITIDO")
    def marcar_no_admitido(self, request, queryset):
        self._cambiar_estado(request, queryset, "no_admitido")

    @admin.action(description="🕒 Marcar como LISTA DE ESPERA")
    def marcar_lista_espera(self, request, queryset):
        self._cambiar_estado(request, queryset, "lista_espera")

    def exportar_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Formación"
        headers = [
            "Fecha", "Usuario", "Convocatoria", "Estado",
            "Email", "Teléfono", "Vínculo con el sector",
            "Nombre", "Apellido", "DNI", "Localidad",
            "Tiene Registro Audiovisual",
        ]
        ws.append(headers)
        qs = queryset.select_related("user", "convocatoria", "persona_humana", "persona_juridica")
        for ins in qs:
            tiene_registro = bool(ins.persona_humana_id or ins.persona_juridica_id)
            ws.append([
                ins.fecha.strftime("%d/%m/%Y %H:%M") if ins.fecha else "",
                ins.user.username,
                ins.convocatoria.titulo if ins.convocatoria else "",
                ins.get_estado_display(),
                ins.user.email or ins.email or "",
                self.contacto_telefono(ins),
                ins.get_vinculo_sector_display() if ins.vinculo_sector else "",
                ins.nombre or "",
                ins.apellido or "",
                ins.dni or "",
                ins.get_localidad_display() if ins.localidad else "",
                "SI" if tiene_registro else "NO",
            ])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inscripciones_formacion.xlsx"'
        wb.save(response)
        return response
    exportar_excel.short_description = "📤 Exportar seleccionadas a Excel (.xlsx)"


# ==========================================
# OBSERVACIÓN ADMINISTRATIVA
# ==========================================

@admin.register(ObservacionFormacion)
class ObservacionFormacionAdmin(admin.ModelAdmin):
    list_display  = ("id", "inscripcion_link", "descripcion", "subsanada", "fecha_creacion")
    list_filter   = ("subsanada",)
    search_fields = ("descripcion", "inscripcion__user__username", "inscripcion__convocatoria__titulo")

    def inscripcion_link(self, obj):
        from django.utils.html import format_html
        url = reverse("admin:formacion_inscripcionformacion_change", args=[obj.inscripcion_id])
        return format_html('<a href="{}">{}</a>', url, str(obj.inscripcion))
    inscripcion_link.short_description = "Inscripción"

    def save_model(self, request, obj, form, change):
        es_nueva = obj.pk is None
        anterior = ObservacionFormacion.objects.filter(pk=obj.pk).first() if not es_nueva else None

        obj.creada_por = obj.creada_por or request.user
        super().save_model(request, obj, form, change)

        # C: marcar inscripción como observada al crear observación pendiente
        if es_nueva and not obj.subsanada:
            insc = obj.inscripcion
            if insc.estado not in ("admitido", "no_admitido"):
                insc.estado = "observado"
                insc.save(update_fields=["estado"])

        if obj.subsanada:
            return

        # Solo enviar si es nueva o cambió descripción
        if anterior and (anterior.descripcion or "").strip() == (obj.descripcion or "").strip():
            return

        inscripcion = obj.inscripcion
        user = getattr(inscripcion, "user", None)
        if not user or not user.email:
            return

        convocatoria_titulo = getattr(getattr(inscripcion, "convocatoria", None), "titulo", "") or ""
        panel_url = request.build_absolute_uri(reverse("usuarios:panel_usuario"))

        asunto = "Observación sobre tu inscripción a formación"
        contexto = {
            "user": user,
            "inscripcion": inscripcion,
            "convocatoria_titulo": convocatoria_titulo,
            "observacion": obj,
            "panel_url": panel_url,
            "anio": timezone.now().year,
        }
        texto = (
            f"Observación sobre tu inscripción.\n\n"
            f"Formación: {convocatoria_titulo or '—'}\n"
            f"Detalle: {obj.descripcion}\n\n"
            f"Ingresá al panel: {panel_url}"
        )

        try:
            html = render_to_string("formacion/email_observacion.html", contexto)
            email = EmailMultiAlternatives(
                subject=asunto,
                body=texto,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                to=[user.email],
            )
            email.attach_alternative(html, "text/html")
            email.send(fail_silently=False)
            messages.success(request, f"Email de observación enviado a {user.email}.")
        except Exception as e:
            messages.error(request, f"No se pudo enviar el email: {e}")
