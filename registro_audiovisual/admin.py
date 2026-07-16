from django.contrib import admin
from django.http import HttpResponse
import openpyxl

from .models import PersonaHumana, PersonaJuridica


# ============================================================
# ACCIÓN GLOBAL — EXPORTAR A EXCEL (.xlsx)
# ============================================================

def exportar_excel(modeladmin, request, queryset):
    """
    Exporta los registros seleccionados a un archivo XLSX.
    Muestra los valores legibles de choices cuando corresponde.
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Campos del modelo
    campos = [field.name for field in modeladmin.model._meta.fields]

    # Encabezados
    for col, campo in enumerate(campos, start=1):
        ws.cell(row=1, column=col, value=campo)

    # Filas
    for row, obj in enumerate(queryset, start=2):
        for col, campo in enumerate(campos, start=1):
            valor = getattr(obj, campo)

            # Si es campo con choices → mostrar display
            try:
                display_method = f"get_{campo}_display"
                if hasattr(obj, display_method):
                    valor = getattr(obj, display_method)()
            except Exception:
                pass

            if valor is None:
                valor = ""

            ws.cell(row=row, column=col, value=str(valor))

    nombre = modeladmin.model.__name__.lower()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}.xlsx"'

    wb.save(response)
    return response


exportar_excel.short_description = "📤 Exportar selección a Excel (.xlsx)"


# ============================================================
# PERSONA HUMANA
# ============================================================

@admin.register(PersonaHumana)
class PersonaHumanaAdmin(admin.ModelAdmin):

    actions = [exportar_excel]
    list_select_related = ("user",)

    list_display = (
        "id",
        "user",
        "fecha_creacion",
        "nombre", "apellido",
        "cuil_cuit",
        "fecha_nacimiento",
        "edad",
        "genero",
        "lugar_residencia",
        "otro_lugar_residencia",
        "nivel_educativo",
        "domicilio_real",
        "codigo_postal_real",
        "telefono",
        "email",
        "situacion_iva",
        "actividad_dgr",
        "domicilio_fiscal",
        "localidad_fiscal",
        "codigo_postal_fiscal",
        "area_desempeno_1",
        "area_desempeno_2",
        "area_cultural",
    )

    search_fields = (
        "nombre", "apellido",
        "cuil_cuit",
        "email",
        "telefono",
    )

    list_filter = (
        "genero",
        "lugar_residencia",
        "nivel_educativo",
        "situacion_iva",
        "actividad_dgr",
        "fecha_creacion",
    )

    ordering = ("-fecha_creacion",)

    readonly_fields = ("fecha_creacion", "fecha_actualizacion", "edad")

    fieldsets = (
        ("🔹 Datos Personales", {
            "fields": (
                "user",
                ("nombre", "apellido"),
                ("cuil_cuit", "fecha_nacimiento", "edad"),
                "genero",
                ("lugar_residencia", "otro_lugar_residencia"),
                "nivel_educativo",
                "domicilio_real",
                "codigo_postal_real",
                ("telefono", "email"),
            )
        }),

        ("🔹 Datos Fiscales", {
            "fields": (
                "situacion_iva",
                "actividad_dgr",
                "domicilio_fiscal",
                "localidad_fiscal",
                "codigo_postal_fiscal",
            )
        }),

        ("🔹 Datos Profesionales", {
            "fields": (
                "area_desempeno_1",
                "area_desempeno_2",
                "area_cultural",
                "portfolio_web",
                "canal_video",
                "instagram",
                "linkedin",
                "link_trabajo_destacado",
            )
        }),

        ("🔹 Metadatos", {
            "fields": ("fecha_creacion", "fecha_actualizacion")
        }),
    )


# ============================================================
# PERSONA JURÍDICA
# ============================================================

@admin.register(PersonaJuridica)
class PersonaJuridicaAdmin(admin.ModelAdmin):

    actions = [exportar_excel]
    list_select_related = ("user",)

    list_display = (
        "id",
        "user",
        "razon_social",
        "nombre_comercial",
        "cuil_cuit",
        "tipo_persona_juridica",
        "domicilio_fiscal",
        "localidad_fiscal",
        "codigo_postal_fiscal",
        "fecha_constitucion",
        "antiguedad",
        "email",
        "telefono",
        "situacion_iva",
        "actividad_dgr",
        "area_desempeno_JJPP_1",
        "area_desempeno_JJPP_2",
        "fecha_creacion",
    )

    search_fields = (
        "razon_social",
        "nombre_comercial",
        "cuil_cuit",
        "email",
        "telefono",
    )

    list_filter = (
        "tipo_persona_juridica",
        "localidad_fiscal",
        "situacion_iva",
        "actividad_dgr",
        "fecha_constitucion",
    )

    ordering = ("-fecha_creacion",)

    readonly_fields = ("fecha_creacion", "fecha_actualizacion", "antiguedad")

    fieldsets = (
        ("🔹 Datos Jurídicos", {
            "fields": (
                "user",
                "tipo_persona_juridica",
                "razon_social",
                "nombre_comercial",
                ("cuil_cuit", "fecha_constitucion", "antiguedad"),
            )
        }),

        ("🔹 Domicilio Fiscal y Contacto", {
            "fields": (
                "domicilio_fiscal",
                "localidad_fiscal",
                "codigo_postal_fiscal",
                ("email", "telefono"),
            )
        }),

        ("🔹 Datos Fiscales", {
            "fields": (
                "situacion_iva",
                "actividad_dgr",
            )
        }),

        ("🔹 Actividad Audiovisual", {
            "fields": (
                "area_desempeno_JJPP_1",
                "area_desempeno_JJPP_2",
                "portfolio_web",
                "canal_video",
                "instagram",
                "linkedin",
                "link_trabajo_destacado",
            )
        }),

        ("🔹 Metadatos", {
            "fields": ("fecha_creacion", "fecha_actualizacion")
        }),
    )
